import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import date
import requests
from bs4 import BeautifulSoup
import json
import re
from datetime import datetime
import unicodedata

from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

# =========================================================
# CONFIGURACIÓN GENERAL
# =========================================================

st.set_page_config(
    page_title="Clasificador Bancario - Grupo Bodeguita Oriente",
    page_icon="🏦",
    layout="wide"
)

# Inicialización de Estados para Evitar NameError y Permitir Acumulación
if "saldo_banesco" not in st.session_state: st.session_state.saldo_banesco = 0.0
if "saldo_bnc" not in st.session_state: st.session_state.saldo_bnc = 0.0
if "saldo_mercantil" not in st.session_state: st.session_state.saldo_mercantil = 0.0
if "saldo_venezuela" not in st.session_state: st.session_state.saldo_venezuela = 0.0
if "saldo_provincial" not in st.session_state: st.session_state.saldo_provincial = 0.0
if "saldo_bancamiga" not in st.session_state: st.session_state.saldo_bancamiga = 0.0
if "saldo_tesoro" not in st.session_state: st.session_state.saldo_tesoro = 0.0
if "saldo_banplus" not in st.session_state: st.session_state.saldo_banplus = 0.0
if "saldo_efectivo" not in st.session_state: st.session_state.saldo_efectivo = 0.0
if "saldo_binance" not in st.session_state: st.session_state.saldo_binance = 0.0
if "total_ingresos_consolidado" not in st.session_state: st.session_state.total_ingresos_consolidado = 0.0
if "total_egresos_ipago_ves" not in st.session_state: st.session_state.total_egresos_ipago_ves = 0.0

# =========================================================
# ESTILOS
# =========================================================

st.markdown("""
<style>

/* Main Background and Fonts */
.stApp {
    background-color: #fafbfc;
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
}

/* Centered Titles and Headers */
h1, h2, h3, h4, h5, h6 {
    text-align: center;
    color: #1e3a5f;
    font-weight: 700;
}

/* Custom premium buttons with hover animations */
.stButton > button {
    background: linear-gradient(135deg, #1e3a5f 0%, #00a8cc 100%);
    color: white;
    border-radius: 8px;
    padding: 12px 28px;
    font-weight: bold;
    border: none;
    box-shadow: 0 4px 15px rgba(0, 168, 204, 0.2);
    transition: all 0.3s ease;
    width: 100%;
}

.stButton > button:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 25px rgba(0, 168, 204, 0.4);
    background: linear-gradient(135deg, #2c5282 0%, #00b5d8 100%);
}

.footer {
    text-align: center;
    color: #a0aec0;
    padding: 24px;
    font-size: 14px;
    border-top: 1px solid #e2e8f0;
    margin-top: 40px;
}

/* Premium KPIs Styles (Vivid Navy to Teal Gradient, Centered) */
.kpi-container {
    display: flex;
    gap: 24px;
    margin-bottom: 35px;
    flex-wrap: wrap;
    justify-content: center;
}
.kpi-card {
    background: linear-gradient(135deg, #1e3a5f 0%, #00b5d8 100%);
    color: white;
    padding: 24px;
    border-radius: 16px;
    box-shadow: 0 10px 30px rgba(0, 181, 216, 0.15);
    flex: 1;
    min-width: 290px;
    max-width: 380px;
    border: 1px solid rgba(255, 255, 255, 0.2);
    transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1);
    text-align: center;
}
.kpi-card:hover {
    transform: translateY(-5px);
    box-shadow: 0 15px 35px rgba(0, 181, 216, 0.3);
}
.kpi-title {
    font-size: 13px;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: #e2e8f0;
    margin-bottom: 8px;
    font-weight: 600;
}
.kpi-value {
    font-size: 28px;
    font-weight: 800;
    color: #ffffff;
    text-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
}
.kpi-subtitle {
    font-size: 12px;
    color: #cbd5e0;
    margin-top: 6px;
    font-style: italic;
}

</style>
""", unsafe_allow_html=True)

# =========================================================
# 🔥 NUEVAS FUNCIONES PARA CONCILIACIÓN MULTIBANCO
# =========================================================

def formato_venezolano(valor):
    """Formatea un número al estilo de moneda venezolana (miles con punto, decimales con coma)"""
    if valor is None:
        return "0,00"
    try:
        parts = f"{float(valor):,.2f}".split(".")
        parts[0] = parts[0].replace(",", ".")
        return ",".join(parts)
    except:
        return "0,00"

def obtener_tasa_bcv(fecha=None, usar_api=False):
    """Obtiene la tasa de la fecha especificada de forma segura"""
    if fecha is None:
        fecha = date.today()
    tasa = obtener_tasa_por_fecha(fecha, usar_api)
    if tasa is None:
        # Intentar obtener la tasa más reciente disponible en nuestro diccionario local
        tasa = 623.0223  # Fallback tasa del 30/06/2026
    return tasa

def obtener_saldo_final_banesco(df_raw):
    """Extrae el saldo final de la columna BALANCE del archivo de Banesco"""
    try:
        df_temp = df_raw.copy()
        if df_temp.shape[1] >= 5:
            balances = df_temp.iloc[:, 4].dropna()
            for val in reversed(balances.values):
                val_clean = convertir_monto(val)
                if val_clean is not None:
                    return val_clean
    except Exception as e:
        st.warning(f"No se pudo extraer el saldo final de Banesco: {e}")
    return 0.0

def obtener_saldo_final_bnc(df_raw, encabezado_idx):
    """Busca la columna de saldo en el reporte de BNC y extrae el último valor numérico"""
    try:
        df_temp = df_raw.iloc[encabezado_idx + 1:].copy()
        headers = df_raw.iloc[encabezado_idx].fillna("").astype(str).str.lower().tolist()
        saldo_col_idx = None
        for idx, h in enumerate(headers):
            if "saldo" in h or "balance" in h:
                saldo_col_idx = idx
                break
        if saldo_col_idx is not None:
            balances = df_temp.iloc[:, saldo_col_idx].dropna()
            for val in reversed(balances.values):
                val_clean = convertir_monto(val)
                if val_clean is not None:
                    return val_clean
    except Exception as e:
        st.warning(f"No se pudo extraer el saldo final de BNC: {e}")
    return 0.0

def obtener_saldo_final_mercantil(df_raw):
    """Extrae el saldo final de la columna BALANCE (columna 9) en Mercantil"""
    try:
        df_temp = df_raw.copy()
        if df_temp.shape[1] >= 9:
            balances = df_temp.iloc[:, 8].dropna()
            for val in reversed(balances.values):
                val_clean = convertir_monto(val)
                if val_clean is not None:
                    return val_clean
    except:
        pass
    return 0.0

def buscar_saldo_en_texto(df_raw):
    """Escanea todo el reporte en busca de celdas con la palabra 'SALDO' o 'DISPONIBLE' y obtiene el número"""
    try:
        # 1. Búsqueda específica de términos de saldo final
        terminos_finales = ["saldo disponible", "saldo actual", "saldo final", "saldo de la cuenta", "monto disponible"]
        for r_idx in range(df_raw.shape[0]):
            for c_idx in range(df_raw.shape[1]):
                val = str(df_raw.iloc[r_idx, c_idx]).lower()
                if any(term in val for term in terminos_finales):
                    # Eliminar fechas del texto antes de buscar números para evitar extraer el año (ej. 2026)
                    texto_sin_fechas = re.sub(r'\b\d{2}[/\-]\d{2}[/\-]\d{4}\b', '', val)
                    texto_sin_fechas = re.sub(r'\b\d{4}[/\-]\d{2}[/\-]\d{2}\b', '', texto_sin_fechas)
                    partes = re.findall(r'[\d\.\,]+', texto_sin_fechas)
                    if partes:
                        for p in reversed(partes):
                            num = convertir_monto(p)
                            if num is not None and num > 100:
                                return num
                    # Celda de la derecha
                    if c_idx + 1 < df_raw.shape[1]:
                        val_right = df_raw.iloc[r_idx, c_idx + 1]
                        num = convertir_monto(val_right)
                        if num is not None and num > 0:
                            return num
                    # Celda de abajo
                    if r_idx + 1 < df_raw.shape[0]:
                        val_below = df_raw.iloc[r_idx + 1, c_idx]
                        num = convertir_monto(val_below)
                        if num is not None and num > 0:
                            return num

        # 2. Búsqueda general si la específica falla (ignora inicial, anterior y promedio)
        for r_idx in range(df_raw.shape[0]):
            for c_idx in range(df_raw.shape[1]):
                val = str(df_raw.iloc[r_idx, c_idx]).lower()
                if ("saldo" in val or "disponible" in val) and "inicial" not in val and "anterior" not in val and "promedio" not in val:
                    # Eliminar fechas del texto antes de buscar números para evitar extraer el año (ej. 2026)
                    texto_sin_fechas = re.sub(r'\b\d{2}[/\-]\d{2}[/\-]\d{4}\b', '', val)
                    texto_sin_fechas = re.sub(r'\b\d{4}[/\-]\d{2}[/\-]\d{2}\b', '', texto_sin_fechas)
                    partes = re.findall(r'[\d\.\,]+', texto_sin_fechas)
                    if partes:
                        for p in reversed(partes):
                            num = convertir_monto(p)
                            if num is not None and num > 100:
                                return num
                    # Celda de la derecha
                    if c_idx + 1 < df_raw.shape[1]:
                        val_right = df_raw.iloc[r_idx, c_idx + 1]
                        num = convertir_monto(val_right)
                        if num is not None and num > 0:
                            return num
                    # Celda de abajo
                    if r_idx + 1 < df_raw.shape[0]:
                        val_below = df_raw.iloc[r_idx + 1, c_idx]
                        num = convertir_monto(val_below)
                        if num is not None and num > 0:
                            return num
    except:
        pass
    return 0.0

def buscar_saldo_inicial(df_raw):
    """Busca un saldo inicial en el texto (Saldo Inicial, Saldo Anterior, etc.)"""
    try:
        for r_idx in range(df_raw.shape[0]):
            for c_idx in range(df_raw.shape[1]):
                val = str(df_raw.iloc[r_idx, c_idx]).lower()
                if "saldo" in val and ("inicial" in val or "anterior" in val):
                    # Celda de la derecha
                    if c_idx + 1 < df_raw.shape[1]:
                        val_right = df_raw.iloc[r_idx, c_idx + 1]
                        num = convertir_monto(val_right)
                        if num is not None and num > 0:
                            return num
                    # Celda de abajo
                    if r_idx + 1 < df_raw.shape[0]:
                        val_below = df_raw.iloc[r_idx + 1, c_idx]
                        num = convertir_monto(val_below)
                        if num is not None and num > 0:
                            return num
    except:
        pass
    return 0.0

def obtener_saldo_final_tesoro(df_raw):
    """Calcula el saldo final de Banco del Tesoro sumando Créditos y Débitos o usando el neto"""
    # Intentar buscar por palabra 'Saldo' primero
    saldo_buscado = buscar_saldo_en_texto(df_raw)
    if saldo_buscado > 0:
        return saldo_buscado
        
    # Intentar buscar saldo inicial
    saldo_inicial = buscar_saldo_inicial(df_raw)
    
    # Si no tiene columna Saldo, sumamos todos los créditos y restamos débitos
    try:
        encabezado = None
        for i in range(min(30, len(df_raw))):
            fila = df_raw.iloc[i].fillna("").astype(str)
            texto = " ".join(fila.tolist()).lower()
            if "fecha" in texto and "referencia" in texto and "concepto" in texto:
                encabezado = i
                break
        if encabezado is None:
            return saldo_inicial
            
        df_temp = df_raw.iloc[encabezado + 1:].copy()
        headers = df_raw.iloc[encabezado].fillna("").astype(str).str.strip().tolist()
        
        col_debito_idx = None
        col_credito_idx = None
        for idx, col in enumerate(headers):
            c_clean = col.lower()
            if "débito" in c_clean or "debito" in c_clean:
                col_debito_idx = idx
            elif "crédito" in c_clean or "credito" in c_clean:
                col_credito_idx = idx
                
        total_creditos = 0.0
        total_debitos = 0.0
        
        if col_credito_idx is not None:
            creditos = df_temp.iloc[:, col_credito_idx].dropna()
            for val in creditos:
                num = convertir_monto(val)
                if num is not None:
                    total_creditos += num
                    
        if col_debito_idx is not None:
            debitos = df_temp.iloc[:, col_debito_idx].dropna()
            for val in debitos:
                num = convertir_monto(val)
                if num is not None:
                    total_debitos += num
                    
        # Retornar: Saldo Inicial + Créditos - Débitos
        saldo_calc = saldo_inicial + total_creditos - total_debitos
        # Si da negativo porque no hay saldo inicial en el archivo, tomar el neto absoluto de los movimientos
        if saldo_calc < 0 and saldo_inicial == 0.0:
            saldo_calc = abs(total_creditos - total_debitos)
        return saldo_calc
    except:
        pass
    return saldo_inicial

def obtener_saldo_final_columna_derecha(df_raw):
    """Busca en las últimas columnas del DataFrame (de derecha a izquierda) el último valor numérico válido"""
    try:
        df_temp = df_raw.copy()
        num_cols = df_temp.shape[1]
        for col_idx in range(num_cols - 1, 3, -1):
            balances = df_temp.iloc[:, col_idx].dropna()
            for val in reversed(balances.values):
                val_clean = convertir_monto(val)
                if val_clean is not None and val_clean > 0:
                    return val_clean
    except:
        pass
    return 0.0

def obtener_saldo_final_bancamiga(df_raw):
    """Busca la columna de saldo en Bancamiga y extrae el último valor numérico"""
    try:
        if isinstance(df_raw.columns, pd.MultiIndex):
            df_raw.columns = df_raw.columns.get_level_values(-1)
        encabezado_idx = None
        for i in range(min(20, len(df_raw))):
            fila = df_raw.iloc[i].fillna("").astype(str)
            texto = " ".join(fila.tolist()).lower()
            if "fecha" in texto and "referencia" in texto and "saldo" in texto:
                encabezado_idx = i
                break
        if encabezado_idx is None:
            return buscar_saldo_en_texto(df_raw)
            
        df_temp = df_raw.iloc[encabezado_idx + 1:].copy()
        headers = df_raw.iloc[encabezado_idx].fillna("").astype(str).str.lower().tolist()
        saldo_col_idx = None
        for idx, h in enumerate(headers):
            if "saldo" in h:
                saldo_col_idx = idx
                break
        if saldo_col_idx is not None:
            balances = df_temp.iloc[:, saldo_col_idx].dropna()
            for val in reversed(balances.values):
                val_clean = convertir_monto(val)
                if val_clean is not None:
                    return val_clean
    except:
        pass
    return buscar_saldo_en_texto(df_raw)

def encontrar_fila_encabezado(df_raw):
    """Busca en las primeras filas una que contenga 'fecha' y ('descripcion' o 'descripción' o 'referencia')"""
    for i in range(min(40, len(df_raw))):
        fila = df_raw.iloc[i].fillna("").astype(str)
        texto = " ".join(fila.tolist()).lower()
        if "fecha" in texto and ("descri" in texto or "referencia" in texto or "monto" in texto):
            return i
    return None

def preparar_df_con_encabezado_dinamico(df_raw):
    """Encuentra el encabezado y limpia el DataFrame para que las columnas tengan los nombres correctos"""
    df_clean = df_raw.copy()
    idx_header = encontrar_fila_encabezado(df_clean)
    if idx_header is not None:
        cols = df_clean.iloc[idx_header].fillna("").astype(str).tolist()
        cols = [c.strip() for c in cols]
        df_clean.columns = cols
        df_clean = df_clean.iloc[idx_header + 1:].reset_index(drop=True)
    return df_clean

def obtener_saldo_final_banplus(df_raw):
    """Extrae el saldo de Banplus buscando 'Saldo Total' al inicio del archivo"""
    try:
        df_temp = df_raw.copy()
        for idx in range(min(15, len(df_temp))):
            for col_idx in range(df_temp.shape[1]):
                val_str = str(df_temp.iloc[idx, col_idx]).strip().lower()
                if "saldo total" in val_str:
                    for r_col in range(col_idx + 1, df_temp.shape[1]):
                        val_saldo = df_temp.iloc[idx, r_col]
                        val_clean = convertir_monto(val_saldo)
                        if val_clean is not None and val_clean > 0:
                            return val_clean
                    if df_temp.shape[1] > 5:
                        val_saldo = df_temp.iloc[idx, 5]
                        val_clean = convertir_monto(val_saldo)
                        if val_clean is not None:
                            return val_clean
    except Exception as e:
        st.warning(f"No se pudo extraer el saldo de Banplus: {e}")
    return 0.0

def obtener_saldo_banco(df_raw, banco, encabezado_idx=None):
    """Obtiene el saldo de un banco combinando extractores específicos y el escáner de texto"""
    if banco == "banesco":
        return obtener_saldo_final_banesco(df_raw) or obtener_saldo_final_columna_derecha(df_raw)
    elif banco == "bnc":
        if encabezado_idx is not None:
            return obtener_saldo_final_bnc(df_raw, encabezado_idx)
        return buscar_saldo_en_texto(df_raw) or obtener_saldo_final_columna_derecha(df_raw)
    elif banco == "mercantil":
        return obtener_saldo_final_mercantil(df_raw) or buscar_saldo_en_texto(df_raw) or obtener_saldo_final_columna_derecha(df_raw)
    elif banco == "tesoro":
        return obtener_saldo_final_tesoro(df_raw)
    elif banco == "bancamiga":
        return obtener_saldo_final_bancamiga(df_raw) or obtener_saldo_final_columna_derecha(df_raw)
    elif banco == "banplus":
        return obtener_saldo_final_banplus(df_raw) or buscar_saldo_en_texto(df_raw) or obtener_saldo_final_columna_derecha(df_raw)
    else:
        return buscar_saldo_en_texto(df_raw) or obtener_saldo_final_columna_derecha(df_raw)

def calcular_saldo_movimientos(df_convertido):
    """Suma los ingresos (NC) y resta los egresos (ND) de un DataFrame en formato unificado"""
    if df_convertido is None or df_convertido.empty:
        return 0.0
    try:
        tipo_col = "TIPO" if "TIPO" in df_convertido.columns else df_convertido.columns[5]
        monto_col = "MONTO BS" if "MONTO BS" in df_convertido.columns else df_convertido.columns[7]
        ingresado = df_convertido[df_convertido[tipo_col] == "NC"][monto_col].sum()
        egresado = df_convertido[df_convertido[tipo_col] == "ND"][monto_col].sum()
        return ingresado - egresado
    except:
        return 0.0

# =========================================================
# 🔥 FUNCIÓN PARA NORMALIZAR TEXTO (eliminar acentos)
# =========================================================

def normalizar_texto(texto):
    """Normaliza texto eliminando acentos y convirtiendo a minúsculas"""
    texto = str(texto)
    texto = (
        unicodedata.normalize("NFKD", texto)
        .encode("ascii", "ignore")
        .decode("utf-8")
    )
    return texto.lower()

# =========================================================
# HEADER
# =========================================================


def leer_excel_sin_encabezados(archivo):
    """Lee archivo Excel sin encabezados detectando el engine correcto"""
    nombre = archivo.name.lower()
    
    try:
        if nombre.endswith('.xls') and not nombre.endswith('.xlsx'):
            try:
                import xlrd
                return pd.read_excel(archivo, sheet_name=0, header=None, engine='xlrd')
            except Exception as e:
                st.warning(f"⚠️ Error leyendo como Excel, intentando como HTML: {str(e)}")
                archivo.seek(0)
                try:
                    tablas = pd.read_html(archivo)
                    if len(tablas) > 0:
                        return tablas[0]
                except Exception:
                    pass
                archivo.seek(0)
                contenido = archivo.read().decode("utf-8", errors="ignore")
                lineas = contenido.split("\n")
                datos = []
                for linea in lineas:
                    if linea.strip():
                        partes = linea.split("\t")
                        if len(partes) == 1:
                            partes = [p for p in linea.split(" ") if p.strip()]
                        if len(partes) > 0:
                            datos.append(partes)
                return pd.DataFrame(datos)
        else:
            return pd.read_excel(archivo, sheet_name=0, header=None, engine='openpyxl')
    except Exception as e:
        st.error(f"No se pudo leer el archivo. Error: {str(e)}")
        st.stop()

def leer_excel_con_encabezados(archivo):
    """Lee archivo Excel con encabezados detectando el engine correcto"""
    nombre = archivo.name.lower()
    
    try:
        if nombre.endswith('.xls') and not nombre.endswith('.xlsx'):
            try:
                import xlrd
                return pd.read_excel(archivo, sheet_name=0, header=0, engine='xlrd')
            except ImportError:
                st.error("❌ Para archivos .xls es necesario instalar xlrd. Ejecuta: pip install xlrd")
                st.stop()
        else:
            return pd.read_excel(archivo, sheet_name=0, header=0, engine='openpyxl')
    except Exception as e:
        try:
            return pd.read_excel(archivo, sheet_name=0, header=None, engine='openpyxl')
        except:
            st.error(f"No se pudo leer el archivo. Error: {str(e)}")
            st.stop()

# =========================================================
# 🔥 DETECCIÓN DE BANCO POR CONTENIDO DEL ARCHIVO
# =========================================================

def detectar_banco_por_contenido(archivo):
    """
    Detecta el banco leyendo el contenido del archivo, no solo el nombre.
    """
    try:
        pos = archivo.tell()
        archivo.seek(0)
        try:
            df_temp = pd.read_excel(archivo, nrows=20, header=None, engine='openpyxl')
            texto = " ".join(df_temp.astype(str).values.flatten()).upper()
            if "BANCAMIGA" in texto or "BANCAMIGA BANCO UNIVERSAL" in texto or "BANCA AMIGA" in texto or "AMIGA" in texto:
                archivo.seek(pos)
                return "bancamiga"
            elif "BANESCO" in texto:
                archivo.seek(pos)
                return "banesco"
            elif "PROVINCIAL" in texto:
                archivo.seek(pos)
                return "provincial"
            elif "BANCO DE VENEZUELA" in texto or "BDV" in texto:
                archivo.seek(pos)
                return "venezuela"
            elif "BNC" in texto:
                archivo.seek(pos)
                return "bnc"
            elif "MERCANTIL" in texto:
                archivo.seek(pos)
                return "mercantil"
            elif "TESORO" in texto or "BANCO DEL TESORO" in texto:
                archivo.seek(pos)
                return "tesoro"
        except Exception:
            pass
        archivo.seek(pos)
        return None
    except Exception:
        return None

def detectar_banco_por_nombre(nombre_archivo):
    """Detecta el banco por el nombre del archivo (fallback)"""
    nombre = nombre_archivo.upper()
    if "TESORO" in nombre or "TESORERIA" in nombre or "TES" in nombre:
        return "tesoro"
    elif "BANCAMIGA" in nombre or "BANCAAMIGA" in nombre or "AMIGA" in nombre:
        return "bancamiga"
    elif "BANESCO" in nombre or re.match(r"^J\d+", nombre_archivo):
        return "banesco"
    elif (
        "MOVIMIENTOS EN MONEDA NACIONAL" in nombre
        or "VENEZUELA" in nombre
        or "BANCO DE VENEZUELA" in nombre
        or "BDV" in nombre
        or "VZLA" in nombre
    ):
        return "venezuela"
    elif "PROVINCIAL" in nombre or "PROV" in nombre:
        return "provincial"
    elif "BNC" in nombre:
        return "bnc"
    elif "MERCANTIL" in nombre:
        return "mercantil"
    return "mercantil"

# =========================================================
# FUNCIONES DE CONVERSIÓN Y LIMPIEZA
# =========================================================

def convertir_monto(valor):
    try:
        if pd.isna(valor):
            return None
        if isinstance(valor, (int, float)):
            numero = float(valor)
            if isinstance(valor, int) and numero >= 100000:
                numero = numero / 100
            return numero
        valor_original = str(valor).strip()
        valor = valor_original
        valor = valor.replace(" ", "").replace("$", "").replace("Bs", "").replace("€", "")
        if valor == "":
            return None
        if "." in valor and "," in valor:
            valor = valor.replace(".", "").replace(",", ".")
        elif "," in valor:
            valor = valor.replace(",", ".")
        numero = float(valor)
        if "." not in valor_original and "," not in valor_original and numero >= 100000:
            numero = numero / 100
        return numero
    except Exception:
        return None

def calcular_usd(monto_bs, tasa):
    try:
        if monto_bs is None or tasa is None or tasa == 0:
            return None
        return round(abs(monto_bs) / abs(tasa), 2)
    except:
        return None

def es_comision(texto, proveedor=None):
    texto = normalizar_texto(texto).strip()
    if proveedor and str(proveedor).strip():
        return False
    if any(x in texto for x in [
        "comisiones sobre servicios contratados",
        "comision gerente comercial",
        "comision vendedor",
        "comision asesor",
        "comision ejecutivo",
        "comision supervisor",
        "comision ventas",
        "comisiones ventas",
        "comision comercial",
        "comisiones comerciales"
    ]):
        return False
    if any(x in texto for x in [
        "pago a proveedores",
        "pago de nomina",
        "nomina",
        "transf entre ctas",
        "transferencia a terceros",
        "pago movil comercial"
    ]):
        return False
    palabras_comision_bancaria = [
        "comision por transferencia",
        "comision pago movil",
        "comisión pago movil",
        "comision x pago de nomina",
        "comision x pago de nominas",
        "itf",
        "impuesto a las transacciones financieras",
        "cargo bancario",
        "mantenimiento de cuenta",
        "comision bancaria",
        "comisión bancaria",
        "cargo por servicio",
        "cargo por transaccion",
        "comision por",
        "comisión por",
        "comision pago movil comercial",
        "comision x pago de nominas mb",
        "com pago otras ctas",
        "com pago otr bcos",
        "comis",
        "comis. cr.i",
        "sms",
        "servicio sms",
        "servicio sms plus",
        "sms plus",
        "domiciliacion j412438905",
        "distribuidora global",
        "emision edo",
        "retencion de impuesto",
        "com. trf",
        "com.serv",
        "emision de estado",
        "below minimum balance charges",
        "stament service",
        "statement service"
    ]
    for patron in palabras_comision_bancaria:
        if patron in texto:
            return True
    return False

# =========================================================
# ENRIQUECER EGRESOS CON IPAGO
# =========================================================

def enriquecer_egresos_con_ipago(df_egresos, df_ipago):
    if df_ipago is None or df_ipago.empty:
        return df_egresos
    df_resultado = df_egresos.copy()
    df_resultado["REFERENCIA_NORM"] = df_resultado["REFERENCIA"].astype(str).str.replace(".0", "", regex=False).str.strip()
    df_ipago["Referencia_Norm"] = df_ipago["Referencia"].astype(str).str.replace(".0", "", regex=False).str.strip()
    
    def generar_variantes_ref(ref):
        ref = str(ref).strip()
        if not ref or ref == "nan":
            return set()
        variantes = set()
        variantes.add(ref)
        ref_sin_ceros = ref.lstrip('0')
        if ref_sin_ceros != ref and ref_sin_ceros:
            variantes.add(ref_sin_ceros)
        if ref.endswith('X'):
            ref_sin_x = ref[:-1]
            if ref_sin_x:
                variantes.add(ref_sin_x)
                ref_sin_x_sin_ceros = ref_sin_x.lstrip('0')
                if ref_sin_x_sin_ceros:
                    variantes.add(ref_sin_x_sin_ceros)
        if 'X' in ref and not ref.endswith('X'):
            ref_sin_x = ref.replace('X', '')
            if ref_sin_x:
                variantes.add(ref_sin_x)
                ref_sin_x_sin_ceros = ref_sin_x.lstrip('0')
                if ref_sin_x_sin_ceros:
                    variantes.add(ref_sin_x_sin_ceros)
        if len(ref) >= 10 and ref.startswith('0'):
            ref_sin_cero_inicial = ref[1:]
            if ref_sin_cero_inicial:
                variantes.add(ref_sin_cero_inicial)
                variantes.add(ref_sin_cero_inicial.lstrip('0'))
        if '.' in ref:
            ref_sin_puntos = ref.replace('.', '')
            if ref_sin_puntos:
                variantes.add(ref_sin_puntos)
                variantes.add(ref_sin_puntos.lstrip('0'))
        return variantes

    ipago_dict = {}
    for _, row in df_ipago.iterrows():
        ref_original = str(row.get("Referencia", "")).strip()
        if not ref_original or ref_original == "nan":
            continue
        variantes = generar_variantes_ref(ref_original)
        for variante in variantes:
            if variante and variante not in ipago_dict:
                ipago_dict[variante] = {
                    "PROVEEDOR": row.get("Proveedor", ""),
                    "TIPO_EGRESO": row.get("Tipo de Egreso", ""),
                    "TIPO_PAGO": row.get("Tipo de Pago", ""),
                    "DESCRIPCION_IPAGO": row.get("Descripción", ""),
                    "FECHA_PAGO": row.get("Fecha Pago", ""),
                    "EMPRESA": row.get("Empresa", ""),
                    "MONTO_IPAGO": row.get("Monto", 0),
                    "MONTO_USD": row.get("Monto USD", 0),
                    "REFERENCIA_ORIGINAL": ref_original
                }

    for idx, row in df_resultado.iterrows():
        ref_banco = str(row.get("REFERENCIA", "")).strip()
        monto_banco = float(row.get("MONTO BS", 0))
        variantes_banco = generar_variantes_ref(ref_banco)
        coincide_ref = False
        datos_encontrados = None
        for variante in variantes_banco:
            if variante in ipago_dict:
                datos_encontrados = ipago_dict[variante]
                coincide_ref = True
                break
        if not coincide_ref:
            for clave, datos in ipago_dict.items():
                monto_ipago = float(datos.get("MONTO_IPAGO", 0))
                if monto_ipago > 0:
                    diferencia = abs(monto_banco - monto_ipago) / max(monto_banco, monto_ipago)
                    if diferencia < 0.01:
                        datos_encontrados = datos
                        coincide_ref = True
                        break
        if datos_encontrados:
            df_resultado.at[idx, "STATUS"] = datos_encontrados["PROVEEDOR"]
            df_resultado.at[idx, "OBSERVACIÓN"] = datos_encontrados["TIPO_EGRESO"]
            df_resultado.at[idx, "TIPO_PAGO"] = datos_encontrados["TIPO_PAGO"]
            df_resultado.at[idx, "PROVEEDOR_IPAGO"] = datos_encontrados["PROVEEDOR"]
            df_resultado.at[idx, "REFERENCIA_IPAGO"] = datos_encontrados.get("REFERENCIA_ORIGINAL", "")
            descripcion_ipago = datos_encontrados["DESCRIPCION_IPAGO"]
            if descripcion_ipago:
                df_resultado.at[idx, "DESCRIPCIÓN"] = descripcion_ipago
                df_resultado.at[idx, "DESCRIPCION_ORIGINAL"] = row.get("DESCRIPCIÓN", "")
            tipo = str(datos_encontrados["TIPO_EGRESO"]).upper()
            desc = str(datos_encontrados["DESCRIPCION_IPAGO"]).upper()
            if "COMISION" in tipo or "COMISION" in desc:
                df_resultado.at[idx, "ES_COMISION"] = True
            else:
                df_resultado.at[idx, "ES_COMISION"] = False
        else:
            df_resultado.at[idx, "STATUS"] = "SIN DATOS IPAGO"
            df_resultado.at[idx, "OBSERVACIÓN"] = "SIN CONCORDANCIA"
            df_resultado.at[idx, "TIPO_PAGO"] = ""
            df_resultado.at[idx, "PROVEEDOR_IPAGO"] = ""
            df_resultado.at[idx, "ES_COMISION"] = False
            
    df_resultado = df_resultado.drop(columns=["REFERENCIA_NORM"], errors="ignore")
    return df_resultado

# =========================================================
# PROCESADORES ESPECÍFICOS POR BANCO
# =========================================================

def procesar_banesco(df):
    st.info("Procesando Banesco...")
    try:
        df.columns = ["FECHA", "REFERENCIA", "DESCRIPCION", "MONTO_RAW", "BALANCE"]
        df.columns = [str(c).strip().upper() for c in df.columns]
        rename_map = {}
        for col in df.columns:
            c = str(col).lower()
            if "fecha" in c: rename_map[col] = "FECHA"
            elif "referencia" in c: rename_map[col] = "REFERENCIA"
            elif "descrip" in c: rename_map[col] = "DESCRIPCION"
            elif "monto" in c: rename_map[col] = "MONTO_RAW"
        df = df.rename(columns=rename_map)
        for col in ["FECHA", "REFERENCIA", "DESCRIPCION", "MONTO_RAW"]:
            if col not in df.columns:
                st.error(f"No existe columna: {col}")
                return pd.DataFrame()
        # Convertir fechas de manera robusta
        def parse_banesco_date(val):
            val_str = str(val).strip()
            if not val_str or val_str == "nan":
                return pd.NaT
            if len(val_str) >= 5 and val_str[:4].isdigit() and val_str[4] in ('/', '-'):
                return pd.to_datetime(val_str, dayfirst=False, errors="coerce")
            return pd.to_datetime(val_str, dayfirst=True, errors="coerce")

        df["FECHA"] = df["FECHA"].apply(parse_banesco_date)
        df = df[df["FECHA"].notna()]
        df["TIPO"] = df["MONTO_RAW"].astype(str).apply(lambda x: "NC" if "+" in x else "ND")
        df["MONTO"] = (df["MONTO_RAW"].astype(str).str.replace("+", "", regex=False)
                       .str.replace("-", "", regex=False)
                       .str.replace(".", "", regex=False)
                       .str.replace(",", ".", regex=False)
                       .str.strip())
        df["MONTO"] = pd.to_numeric(df["MONTO"], errors="coerce")
        df = df[df["MONTO"].notna()]
        df = df[df["MONTO"] > 0]
        df = df[["FECHA", "REFERENCIA", "DESCRIPCION", "TIPO", "MONTO"]]
        st.success(f"Banesco OK: {len(df)} movimientos")
        return df
    except Exception as e:
        st.error(f"Error Banesco: {str(e)}")
        return pd.DataFrame()

def procesar_provincial(df):
    st.info("🔍 Procesando archivo de Provincial...")
    try:
        encabezado_idx = None
        for i in range(min(30, len(df))):
            fila = df.iloc[i]
            fila_str = [str(val) for val in fila.tolist()]
            texto_fila = " ".join(fila_str).upper()
            if "CONCEPTO" in texto_fila and "IMPORTE" in texto_fila:
                encabezado_idx = i
                break
        if encabezado_idx is None:
            st.error("❌ No se encontró la fila de encabezados en el archivo Provincial.")
            return pd.DataFrame()
        headers = df.iloc[encabezado_idx].astype(str).str.strip().tolist()
        rename_map = {}
        for col in headers:
            col_clean = str(col).strip().upper()
            if "OPERAC" in col_clean or "FECHA" in col_clean: rename_map[col] = "FECHA"
            elif "F. VALOR" in col_clean: rename_map[col] = "FECHA_VALOR"
            elif "CÓDIGO" in col_clean or "CODIGO" in col_clean: rename_map[col] = "CODIGO"
            elif "Nº. DOC" in col_clean or "NRO DOC" in col_clean or "DOC" in col_clean: rename_map[col] = "REFERENCIA"
            elif "CONCEPTO" in col_clean: rename_map[col] = "DESCRIPCION"
            elif "IMPORTE" in col_clean: rename_map[col] = "MONTO"
        df.columns = headers
        df = df.iloc[encabezado_idx + 1:].reset_index(drop=True)
        df = df.rename(columns=rename_map)
        if "FECHA" in df.columns:
            df["FECHA"] = df["FECHA"].astype(str).str.strip()
            df = df[~df["FECHA"].str.contains("FECHA|SALDO|Período", case=False, na=False)]
            df = df[df["FECHA"].str.match(r'^\d{2}[-/]\d{2}[-/]\d{2,4}$', na=False)]
            df["FECHA"] = pd.to_datetime(df["FECHA"], dayfirst=True, errors="coerce")
            df = df[df["FECHA"].notna()]
        else:
            return pd.DataFrame()
        if "MONTO" in df.columns:
            df["MONTO"] = df["MONTO"].astype(str).str.replace(" ", "", regex=False).str.replace(".", "", regex=False).str.replace(",", ".", regex=False).str.replace("'", "", regex=False)
            df["MONTO"] = pd.to_numeric(df["MONTO"], errors="coerce")
            df = df[df["MONTO"].notna()]
            df["TIPO"] = df["MONTO"].apply(lambda x: "NC" if x > 0 else "ND" if x < 0 else "")
            df["MONTO"] = df["MONTO"].abs()
            df = df[df["MONTO"] > 0]
        else:
            return pd.DataFrame()
        if "REFERENCIA" not in df.columns: df["REFERENCIA"] = ""
        else: df["REFERENCIA"] = df["REFERENCIA"].astype(str).str.strip().str.replace("'", "", regex=False)
        if "DESCRIPCION" not in df.columns: df["DESCRIPCION"] = ""
        else: df["DESCRIPCION"] = df["DESCRIPCION"].astype(str).str.strip()
        df["ES_COMISION"] = df["DESCRIPCION"].str.contains("COMIS", case=False, na=False)
        df_resultado = df[["FECHA", "REFERENCIA", "DESCRIPCION", "TIPO", "MONTO", "ES_COMISION"]].copy()
        st.success(f"✅ Provincial OK: {len(df_resultado)} movimientos")
        return df_resultado
    except Exception as e:
        st.error(f"❌ Error procesando Provincial: {str(e)}")
        return pd.DataFrame()

def procesar_bnc(df):
    st.info("Procesando archivo BNC...")
    try:
        encabezado = None
        for i in range(min(30, len(df))):
            fila = df.iloc[i].fillna("").astype(str)
            texto = " ".join(fila.tolist()).lower()
            if "fecha" in texto and ("descripcion" in texto or "descripción" in texto):
                encabezado = i
                break
        if encabezado is None:
            return pd.DataFrame()
        headers = []
        for idx, col in enumerate(df.iloc[encabezado]):
            col = str(col).strip().replace("\n", " ")
            if col == "" or col.lower() == "nan": col = f"COLUMNA_{idx}"
            headers.append(col)
        headers_unicos = []
        contador = {}
        for h in headers:
            if h in contador:
                contador[h] += 1
                nuevo = f"{h}_{contador[h]}"
            else:
                contador[h] = 0
                nuevo = h
            headers_unicos.append(nuevo)
        df.columns = headers_unicos
        df = df.iloc[encabezado + 1:].reset_index(drop=True)
        rename_map = {}
        for col in df.columns:
            col_str = str(col).strip().lower()
            if "fecha" in col_str: rename_map[col] = "FECHA"
            elif "descripcion" in col_str or "descripción" in col_str or "concepto" in col_str: rename_map[col] = "DESCRIPCION"
            elif "referencia" in col_str: rename_map[col] = "REFERENCIA"
            elif "credito" in col_str or "haber" in col_str: rename_map[col] = "CREDITO"
            elif "debito" in col_str or "debe" in col_str: rename_map[col] = "DEBITO"
        df = df.rename(columns=rename_map)
        if "FECHA" in df.columns:
            df["FECHA"] = pd.to_datetime(df["FECHA"], dayfirst=True, errors="coerce")
            df = df[df["FECHA"].notna()]
        df["CREDITO"] = pd.to_numeric(df.get("CREDITO", 0), errors="coerce").fillna(0) if "CREDITO" in df.columns else 0
        df["DEBITO"] = pd.to_numeric(df.get("DEBITO", 0), errors="coerce").fillna(0) if "DEBITO" in df.columns else 0
        df["MONTO"] = df["CREDITO"] - df["DEBITO"]
        df["TIPO"] = df["MONTO"].apply(lambda x: "NC" if x > 0 else "ND")
        df["MONTO"] = df["MONTO"].abs()
        df = df[df["MONTO"] != 0]
        st.success(f"Registros BNC OK: {len(df)}")
        return df
    except Exception as e:
        st.error(f"Error BNC: {e}")
        return pd.DataFrame()

def procesar_tesoro(df):
    st.info("Procesando Banco del Tesoro...")
    try:
        encabezado = None
        for i in range(min(20, len(df))):
            fila = df.iloc[i].astype(str)
            texto = " ".join(map(str, fila.tolist())).lower()
            if "fecha" in texto and "referencia" in texto and "concepto" in texto:
                encabezado = i
                break
        if encabezado is None:
            return pd.DataFrame()
        df.columns = df.iloc[encabezado]
        df = df.iloc[encabezado + 1:].reset_index(drop=True)
        df.columns = [str(c).strip() for c in df.columns]
        rename_map = {}
        for col in df.columns:
            c = str(col).strip().lower()
            if "fecha" in c: rename_map[col] = "FECHA"
            elif "referencia" in c: rename_map[col] = "REFERENCIA"
            elif "concepto" in c: rename_map[col] = "DESCRIPCION"
            elif "débito" in c or "debito" in c: rename_map[col] = "DEBITO"
            elif "crédito" in c or "credito" in c: rename_map[col] = "CREDITO"
            elif "código" in c or "codigo" in c: rename_map[col] = "TIPO"
        df = df.rename(columns=rename_map)
        if "FECHA" not in df.columns: return pd.DataFrame()
        df["FECHA"] = pd.to_datetime(df["FECHA"], dayfirst=True, errors="coerce")
        df = df[df["FECHA"].notna()]
        def limpiar_numero(valor):
            valor = str(valor).replace(".", "").replace(",", ".")
            try: return float(valor)
            except: return 0
        df["CREDITO"] = df.get("CREDITO", 0).apply(limpiar_numero) if "CREDITO" in df.columns else 0
        df["DEBITO"] = df.get("DEBITO", 0).apply(limpiar_numero) if "DEBITO" in df.columns else 0
        df["MONTO"] = df["CREDITO"] - df["DEBITO"]
        df["TIPO"] = df["MONTO"].apply(lambda x: "NC" if x > 0 else "ND")
        df["MONTO"] = df["MONTO"].abs()
        df = df[df["MONTO"] > 0]
        df = df[["FECHA", "REFERENCIA", "DESCRIPCION", "TIPO", "MONTO"]]
        st.success(f"Tesoro OK: {len(df)} registros")
        return df
    except Exception as e:
        st.error(f"Error Tesoro: {str(e)}")
        return pd.DataFrame()

def procesar_bancamiga(df):
    st.info("🔍 Procesando archivo de Bancamiga...")
    try:
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(-1)
        columnas = [str(c).strip().upper() for c in df.columns]
        if "FECHA" in columnas and "REFERENCIA" in columnas:
            rename_map = {
                "NRO.": "NRO", "NRO": "NRO", "FECHA": "FECHA", "REFERENCIA": "REFERENCIA",
                "CONCEPTO": "DESCRIPCION", "DÉBITO": "DEBITO", "DEBITO": "DEBITO",
                "CRÉDITO": "CREDITO", "CREDITO": "CREDITO", "SALDO": "SALDO"
            }
            df.columns = [rename_map.get(str(c).strip().upper(), str(c).strip().upper()) for c in df.columns]
        else:
            encabezado_idx = None
            for i in range(min(30, len(df))):
                fila = df.iloc[i]
                fila_str = [str(val) for val in fila.tolist()]
                texto_fila = " ".join(fila_str).upper()
                if "NRO" in texto_fila and "FECHA" in texto_fila and "REFERENCIA" in texto_fila:
                    encabezado_idx = i
                    break
            if encabezado_idx is None: return pd.DataFrame()
            headers = df.iloc[encabezado_idx].astype(str).str.strip().tolist()
            rename_map = {}
            for col in headers:
                col_clean = str(col).strip().upper()
                if "NRO" in col_clean or "Nº" in col_clean: rename_map[col] = "NRO"
                elif "FECHA" in col_clean: rename_map[col] = "FECHA"
                elif "REFERENCIA" in col_clean: rename_map[col] = "REFERENCIA"
                elif "CONCEPTO" in col_clean: rename_map[col] = "DESCRIPCION"
                elif "DÉBITO" in col_clean or "DEBITO" in col_clean: rename_map[col] = "DEBITO"
                elif "CRÉDITO" in col_clean or "CREDITO" in col_clean: rename_map[col] = "CREDITO"
                elif "SALDO" in col_clean: rename_map[col] = "SALDO"
            df.columns = headers
            df = df.iloc[encabezado_idx + 1:].reset_index(drop=True)
            df = df.rename(columns=rename_map)
            
        if "FECHA" not in df.columns: return pd.DataFrame()
        # Filtrar filas que no son movimientos
        fechas_str_col = df["FECHA"].astype(str).str.strip()
        df = df[~fechas_str_col.str.contains("FECHA|SALDO|TOTAL|CRÉDITO|CREDITO|DÉBITO|DEBITO", case=False, na=False)]
        
        # Convertir fechas de manera robusta
        df["FECHA_DT"] = pd.to_datetime(df["FECHA"], dayfirst=True, errors="coerce")
        mask = df["FECHA_DT"].isna()
        if mask.any():
            df.loc[mask, "FECHA_DT"] = pd.to_datetime(df.loc[mask, "FECHA"].astype(str).str.strip(), dayfirst=True, errors="coerce")
        df["FECHA"] = df["FECHA_DT"]
        df = df[df["FECHA"].notna()]
        
        def limpiar_monto(val):
            val_str = str(val).strip().replace(" ", "")
            if not val_str or val_str == "nan":
                return 0.0
            if "," in val_str:
                val_str = val_str.replace(".", "").replace(",", ".")
            return pd.to_numeric(val_str, errors="coerce")

        df["DEBITO"] = df["DEBITO"].apply(limpiar_monto).fillna(0) if "DEBITO" in df.columns else 0.0
        df["CREDITO"] = df["CREDITO"].apply(limpiar_monto).fillna(0) if "CREDITO" in df.columns else 0.0
        
        df["MONTO"] = df["CREDITO"] - df["DEBITO"]
        df["TIPO"] = df["MONTO"].apply(lambda x: "NC" if x > 0 else "ND" if x < 0 else "")
        df["MONTO"] = df["MONTO"].abs()
        df = df[df["MONTO"] > 0]
        if "REFERENCIA" not in df.columns: df["REFERENCIA"] = ""
        else: df["REFERENCIA"] = df["REFERENCIA"].astype(str).str.strip().str.replace("'", "", regex=False)
        if "DESCRIPCION" not in df.columns: df["DESCRIPCION"] = ""
        else: df["DESCRIPCION"] = df["DESCRIPCION"].astype(str).str.strip()
        df["ES_COMISION"] = df["DESCRIPCION"].str.contains("Comisi", case=False, na=False)
        df_resultado = df[["FECHA", "REFERENCIA", "DESCRIPCION", "TIPO", "MONTO", "ES_COMISION"]].copy()
        st.success(f"✅ Bancamiga OK: {len(df_resultado)} movimientos")
        return df_resultado
    except Exception as e:
        st.error(f"❌ Error Bancamiga: {str(e)}")
        return pd.DataFrame()

def procesar_banplus(df):
    st.info("🔍 Procesando archivo de BanPlus...")
    try:
        columnas = [str(c).strip().upper() for c in df.columns]
        if "FECHA" in columnas and "REFERENCIA" in columnas:
            rename_map = {
                "FECHA": "FECHA", "REFERENCIA": "REFERENCIA", "DESCRIPCION": "DESCRIPCION",
                "CONCEPTO": "DESCRIPCION", "DEBITO": "DEBITO", "DEBITOS": "DEBITO",
                "DÉBITO": "DEBITO", "CREDITO": "CREDITO", "CREDITOS": "CREDITO",
                "CRÉDITO": "CREDITO", "SALDO": "SALDO"
            }
            df.columns = [rename_map.get(str(c).strip().upper(), str(c).strip().upper()) for c in df.columns]
        else:
            encabezado_idx = None
            for i in range(min(30, len(df))):
                fila = df.iloc[i]
                fila_str = [str(val) for val in fila.tolist()]
                texto_fila = " ".join(fila_str).upper()
                if "FECHA" in texto_fila and ("REFERENCIA" in texto_fila or "REF" in texto_fila or "DESCRIPCION" in texto_fila or "CONCEPTO" in texto_fila):
                    encabezado_idx = i
                    break
            if encabezado_idx is None: return pd.DataFrame()
            headers = df.iloc[encabezado_idx].astype(str).str.strip().tolist()
            rename_map = {}
            for col in headers:
                col_clean = str(col).strip().upper()
                if "FECHA" in col_clean: rename_map[col] = "FECHA"
                elif "REFERENCIA" in col_clean or "REF" in col_clean: rename_map[col] = "REFERENCIA"
                elif "CONCEPTO" in col_clean or "DESCRIP" in col_clean: rename_map[col] = "DESCRIPCION"
                elif "DÉBITO" in col_clean or "DEBITO" in col_clean or "EGRESO" in col_clean or "CARGO" in col_clean: rename_map[col] = "DEBITO"
                elif "CRÉDITO" in col_clean or "CREDITO" in col_clean or "INGRESO" in col_clean or "ABONO" in col_clean: rename_map[col] = "CREDITO"
                elif "SALDO" in col_clean: rename_map[col] = "SALDO"
            df.columns = headers
            df = df.iloc[encabezado_idx + 1:].reset_index(drop=True)
            df = df.rename(columns=rename_map)
            
        if "FECHA" not in df.columns: return pd.DataFrame()
        df["FECHA"] = df["FECHA"].astype(str).str.strip()
        df = df[~df["FECHA"].str.contains("FECHA|SALDO|TOTAL|CRÉDITO|CREDITO|DÉBITO|DEBITO", case=False, na=False)]
        df["FECHA"] = pd.to_datetime(df["FECHA"], format="%d/%m/%Y", errors="coerce")
        mask = df["FECHA"].isna()
        if mask.any():
            df.loc[mask, "FECHA"] = pd.to_datetime(df.loc[mask, "FECHA"].astype(str), dayfirst=True, errors="coerce")
        df = df[df["FECHA"].notna()]
        
        df["DEBITO"] = df["DEBITO"].apply(mono_limpiar_monto_banplus) if "DEBITO" in df.columns else 0.0
        df["CREDITO"] = df["CREDITO"].apply(mono_limpiar_monto_banplus) if "CREDITO" in df.columns else 0.0
        
        df["MONTO"] = df["CREDITO"] - df["DEBITO"]
        df["TIPO"] = df["MONTO"].apply(lambda x: "NC" if x > 0 else "ND" if x < 0 else "")
        df["MONTO"] = df["MONTO"].abs()
        df = df[df["MONTO"] > 0]
        if "REFERENCIA" not in df.columns: df["REFERENCIA"] = ""
        else: df["REFERENCIA"] = df["REFERENCIA"].astype(str).str.strip().str.replace("'", "", regex=False)
        if "DESCRIPCION" not in df.columns: df["DESCRIPCION"] = ""
        else: df["DESCRIPCION"] = df["DESCRIPCION"].astype(str).str.strip()
        df["ES_COMISION"] = df["DESCRIPCION"].str.contains("Comisi|sms|servicio sms|sms plus", case=False, na=False)
        df_resultado = df[["FECHA", "REFERENCIA", "DESCRIPCION", "TIPO", "MONTO", "ES_COMISION"]].copy()
        st.success(f"✅ BanPlus OK: {len(df_resultado)} movimientos")
        return df_resultado
    except Exception as e:
        st.error(f"❌ Error BanPlus: {str(e)}")
        return pd.DataFrame()

def procesar_venezuela_simple(df):
    st.info("🔍 Procesando Banco de Venezuela (MODO SIMPLE)...")
    try:
        col_fecha = 3
        col_ref = 1
        col_desc = 2
        col_tipo = 4
        col_credito = 5
        col_debito = 6
        
        movimientos = []
        for idx in range(1, len(df)):
            try:
                fila = df.iloc[idx]
                if pd.isna(fila[col_fecha]): continue
                fecha_raw = str(fila[col_fecha]).strip()
                fecha_val = pd.to_datetime(fecha_raw, format="%d/%m/%Y", errors="coerce")
                if pd.isna(fecha_val):
                    fecha_val = pd.to_datetime(fecha_raw, dayfirst=True, errors="coerce")
                if pd.isna(fecha_val): continue
                
                referencia = str(fila[col_ref]).strip() if pd.notna(fila[col_ref]) else ""
                descripcion = str(fila[col_desc]).strip() if pd.notna(fila[col_desc]) else ""
                tipo_mov = str(fila[col_tipo]).strip().upper() if pd.notna(fila[col_tipo]) else ""
                
                desc_upper = descripcion.upper()
                if desc_upper in ["SALDO INICIAL", "SALDO FINAL", "TOTALES"]: continue
                
                val_credito = 0
                if pd.notna(fila[col_credito]):
                    clean_cred = str(fila[col_credito]).strip().replace(".", "").replace(",", ".")
                    try: val_credito = float(clean_cred)
                    except: pass
                
                val_debito = 0
                if pd.notna(fila[col_debito]):
                    clean_deb = str(fila[col_debito]).strip().replace(".", "").replace(",", ".")
                    try: val_debito = float(clean_deb)
                    except: pass
                
                monto = 0
                tipo = ""
                if tipo_mov == "NC":
                    monto = abs(val_credito)
                    tipo = "NC"
                elif tipo_mov == "ND":
                    monto = abs(val_debito)
                    tipo = "ND"
                else:
                    if abs(val_credito) > 0:
                        monto = abs(val_credito)
                        tipo = "NC"
                    elif abs(val_debito) > 0:
                        monto = abs(val_debito)
                        tipo = "ND"
                    else: continue
                
                if monto <= 0: continue
                movimientos.append({
                    "FECHA": fecha_val.strftime("%d/%m/%Y"),
                    "FECHA_OBJ": fecha_val,
                    "REFERENCIA": referencia,
                    "DESCRIPCION": descripcion,
                    "TIPO": tipo,
                    "MONTO": monto
                })
            except:
                continue
        df_resultado = pd.DataFrame(movimientos)
        st.success(f"✅ Venezuela OK: {len(df_resultado)} movimientos")
        return df_resultado
    except Exception as e:
        st.error(f"❌ Error en BDV: {str(e)}")
        return pd.DataFrame()

# =========================================================
# CONVERTIDORES A FORMATO MERCANTIL
# =========================================================

def convertir_venezuela_a_formato_mercantil(df):
    datos_convertidos = []
    for idx, fila in df.iterrows():
        try:
            fecha = fila["FECHA_OBJ"] if "FECHA_OBJ" in fila else fila["FECHA"]
            if pd.isna(fecha): continue
            if isinstance(fecha, (pd.Timestamp, datetime)):
                fecha_str = fecha.strftime("%d/%m/%Y")
            else:
                fecha_str = str(fecha)
            
            tipo = fila.get("TIPO", "") or ""
            descripcion = fila.get("DESCRIPCION", "") or ""
            referencia = fila.get("REFERENCIA", "") or ""
            monto = fila.get("MONTO", 0) or 0
            
            fila_convertida = ["", "", "", fecha_str, referencia, tipo, descripcion, monto, "", False]
            datos_convertidos.append(fila_convertida)
        except:
            continue
    df_convertido = pd.DataFrame(datos_convertidos)
    return df_convertido if len(df_convertido) > 0 else pd.DataFrame()

def convertir_a_formato_mercantil(df, banco):
    datos_convertidos = []
    df_temp = df.copy()
    
    # Si es Mercantil o no hay columnas de texto esperadas, mapeamos por índice
    if banco == "mercantil" or "FECHA" not in [str(c).strip().upper() for c in df_temp.columns]:
        for idx, fila in df_temp.iterrows():
            try:
                if len(fila) < 8: continue
                fecha = fila.iloc[3]
                if pd.isna(fecha): continue
                fecha_str = str(fecha).strip().replace(".0", "")
                
                tipo = str(fila.iloc[5]).strip()
                descripcion = str(fila.iloc[6]).strip()
                referencia = str(fila.iloc[4]).strip()
                monto = fila.iloc[7]
                es_comision_flag = bool(fila.iloc[9]) if len(fila) > 9 else False
                
                fila_convertida = ["", "", "", fecha_str, referencia, tipo, descripcion, monto, "", es_comision_flag]
                datos_convertidos.append(fila_convertida)
            except:
                continue
    else:
        df_temp.columns = [str(c).strip().upper() for c in df_temp.columns]
        for idx, fila in df_temp.iterrows():
            try:
                fecha = fila.get("FECHA", "")
                if pd.isna(fecha): continue
                if isinstance(fecha, (pd.Timestamp, datetime)):
                    fecha_str = fecha.strftime("%d/%m/%Y")
                else:
                    fecha_str = str(fecha)
                
                tipo = fila.get("TIPO", "") or ""
                descripcion = fila.get("DESCRIPCION", "") or ""
                referencia = fila.get("REFERENCIA", "") or ""
                monto = fila.get("MONTO", 0) or 0
                es_comision_flag = bool(fila.get("ES_COMISION", False))
                
                fila_convertida = ["", "", "", fecha_str, referencia, tipo, descripcion, monto, "", es_comision_flag]
                datos_convertidos.append(fila_convertida)
            except:
                continue
                
    df_convertido = pd.DataFrame(datos_convertidos)
    return df_convertido if len(df_convertido) > 0 else pd.DataFrame()

# =========================================================
# OBTENER TASA BCV HISTORICA LOCAL
# =========================================================

@st.cache_data(ttl=3600)
def obtener_tasa_bcv_fecha(fecha_obj):
    tasas_bcv_local = {
        "01/06/2026": 554.4258, "02/06/2026": 557.9741, "03/06/2026": 558.6436,
        "04/06/2026": 560.3753, "05/06/2026": 563.2892, "06/06/2026": 567.6828,
        "07/06/2026": 567.6828, "08/06/2026": 567.6828, "09/06/2026": 567.6828,
        "10/06/2026": 572.6784, "11/06/2026": 577.5461, "12/06/2026": 582.6862,
        "13/06/2026": 587.4059, "14/06/2026": 587.4059, "15/06/2026": 587.4059,
        "16/06/2026": 592.5163, "17/06/2026": 596.7824, "18/06/2026": 602.3324,
        "19/06/2026": 607.3919, "20/06/2026": 612.4332, "21/06/2026": 612.4332,
        "22/06/2026": 612.4332, "23/06/2026": 617.6388, "24/06/2026": 621.5299,
        "25/06/2026": 621.5299, "26/06/2026": 622.2135, "27/06/2026": 623.0223,
        "28/06/2026": 623.0223, "29/06/2026": 623.0223, "30/06/2026": 623.0223,
    }
    fecha_str = fecha_obj.strftime("%d/%m/%Y")
    return tasas_bcv_local.get(fecha_str, None)

def obtener_tasa_por_fecha(fecha_obj, usar_api=False):
    return obtener_tasa_bcv_fecha(fecha_obj)

# =========================================================
# 🔥 PROCESAMIENTO PRINCIPAL - CLASIFICACIÓN
# =========================================================

def procesar_archivo(df, usar_api=False, banco=""):
    ingresos = []
    egresos = []
    comisiones = []
    registros_procesados = set()
    
    tipos_ingresos = ["NC", "C", "CREDITO", "ABONO", "DP", "DEP"]
    tipos_egresos = ["ND", "D", "DEBITO", "DEBIT"]
    cache_tasas = {}

    for _, fila in df.iterrows():
        try:
            if len(fila) < 10: continue
            fecha_raw = str(fila[3]).strip()
            if fecha_raw.lower() == "nan": continue
            fecha_raw = fecha_raw.replace(".0", "")
            if len(fecha_raw) == 7 and fecha_raw.isdigit():
                fecha = f"0{fecha_raw[0]}/{fecha_raw[1:3]}/{fecha_raw[3:]}"
            elif len(fecha_raw) == 8 and fecha_raw.isdigit():
                fecha = f"{fecha_raw[0:2]}/{fecha_raw[2:4]}/{fecha_raw[4:]}"
            else:
                fecha = fecha_raw

            tipo = str(fila[5]).strip().upper()
            descripcion = str(fila[6]).strip()
            referencia = str(fila[4]).strip()
            monto_bs = convertir_monto(fila[7])
            if monto_bs is None or monto_bs == 0: continue
            
            fecha_obj = pd.to_datetime(fecha, dayfirst=True, errors="coerce")
            if pd.isna(fecha_obj): continue
            
            fecha_key = fecha_obj.strftime("%d/%m/%Y")
            tasa = cache_tasas.get(fecha_key) or obtener_tasa_por_fecha(fecha_obj, usar_api) or 1.0
            cache_tasas[fecha_key] = tasa

            monto_usd = calcular_usd(monto_bs, tasa)
            if monto_usd is None: continue
            
            texto = descripcion.upper()
            if texto in ["SALDO", "DESCRIPCION", "DESCRIPCIÓN", "REFERENCIA", "MOVIMIENTO", "FECHA", "SALDO INICIAL", "SALDO FINAL"]:
                continue

            registro = {
                "FECHA": fecha, "REFERENCIA": referencia, "DESCRIPCIÓN": descripcion,
                "MONTO BS": round(abs(monto_bs), 2), "TASA BCV": round(tasa, 4), "MONTO USD": monto_usd,
                "STATUS": "", "OBSERVACIÓN": "", "TIPO_PAGO": "", "PROVEEDOR_IPAGO": "", "DESCRIPCION_ORIGINAL": ""
            }

            clave = (fecha, referencia, descripcion, monto_usd, tipo)
            if clave in registros_procesados: continue
            registros_procesados.add(clave)

            # Reglas específicas por banco
            es_comision_banco = False
            if banco == "provincial" and "COMIS" in texto: es_comision_banco = True
            elif banco == "bancamiga" and "COMISI" in texto: es_comision_banco = True
            elif banco == "mercantil":
                patrones = [
                    "OP.CRED.DIRT. CLTE-CLTE", "OP CRED DIRT CLTE CLTE", "COMISION PAGO MOVIL COMERCIAL",
                    "COMISION POR TRANSFERENCIA DE FONDOS", "COMISION X PAGO DE NOMINAS",
                    "COMISION PAGO MOVIL COMERCIAL INTERBANCARIO", "COMISION X PAGO DE NOMINAS MB",
                    "ITF", "IMPUESTO A LAS TRANSACCIONES FINANCIERAS", "CARGO BANCARIO",
                    "MANTENIMIENTO DE CUENTA", "COMISION POR TRANSFERENCIA"
                ]
                if any(p in texto for p in patrones): es_comision_banco = True
            
            # BDV / Venezuela comisiones por descripción o referencia
            if banco == "venezuela":
                patrones_bdv = [
                    "COM PAGO OTRAS CTAS", "COMISION PAGO A PROVEEDORES", "COM PAGO OTR BCOS",
                    "COM PAGO OTRAS CTAS JUR NAT", "COM PAGO OTRAS CTAS JUR JUR", "COMISION POR TRANSFERENCIA",
                    "COMISION PAGO MOVIL", "COMISIÓN PAGO MOVIL", "COMISION X PAGO DE NOMINA",
                    "COMISION X PAGO DE NOMINAS", "ITF", "IMPUESTO A LAS TRANSACCIONES FINANCIERAS",
                    "CARGO BANCARIO", "MANTENIMIENTO DE CUENTA", "COMISION BANCARIA", "COMISIÓN BANCARIA",
                    "CARGO POR SERVICIO", "CARGO POR TRANSACCION", "COMISION PAGO MOVIL COMERCIAL",
                    "COMISION X PAGO DE NOMINAS MB", "DOMICILIACION J412438905", "DISTRIBUIDORA GLOBAL",
                    "DOMICILIACION"
                ]
                if any(p in texto for p in patrones_bdv) or referencia.startswith(("970", "972", "067")) or (tipo == "ND" and "COM" in texto):
                    es_comision_banco = True
            elif banco == "tesoro":
                patrones_tesoro = [
                    "BELOW MINIMUM BALANCE CHARGES", "STAMENT SERVICE", "STATEMENT SERVICE",
                    "COMIS", "COMISION", "CARGO BANCARIO", "CARGO POR SERVICIO"
                ]
                if any(p in texto for p in patrones_tesoro):
                    es_comision_banco = True

            if es_comision_banco or es_comision(descripcion):
                comisiones.append(registro)
            elif tipo in tipos_ingresos:
                ingresos.append(registro)
            elif tipo in tipos_egresos:
                egresos.append(registro)
        except:
            continue
    return ingresos, egresos, comisiones

# =========================================================
# INTERFAZ PRINCIPAL - EJECUCIÓN
# =========================================================


# =========================================================
# 🔥 MONOBANCO FUNCTIONS (NAMESPACED TO AVOID OVERWRITING)
# =========================================================

def mono_leer_excel_sin_encabezados(archivo):
    """Lee archivo Excel sin encabezados detectando el engine correcto"""
    nombre = archivo.name.lower()
    
    try:
        if nombre.endswith('.xls') and not nombre.endswith('.xlsx'):
            try:
                import xlrd
                # Intentar leer con xlrd
                return pd.read_excel(archivo, sheet_name=0, header=None, engine='xlrd')
            except Exception as e:
                # Si falla, intentar leer como HTML o texto
                print(f"[Reader Warning] Error leyendo como Excel, intentando como HTML: {str(e)}")
                
                archivo.seek(0)
                
                try:
                    tablas = pd.read_html(archivo)
                    
                    if len(tablas) > 0:
                        return tablas[0]
                    
                except Exception:
                    pass
                
                archivo.seek(0)
                
                contenido = archivo.read().decode("utf-8", errors="ignore")
                
                lineas = contenido.split("\n")
                
                datos = []
                
                for linea in lineas:
                    
                    if linea.strip():
                        
                        partes = linea.split("\t")
                        
                        if len(partes) == 1:
                            
                            partes = [p for p in linea.split(" ") if p.strip()]
                        
                        if len(partes) > 0:
                            
                            datos.append(partes)
                
                return pd.DataFrame(datos)
        else:
            return pd.read_excel(archivo, sheet_name=0, header=None, engine='openpyxl')
    except Exception as e:
        st.error(f"No se pudo leer el archivo. Error: {str(e)}")
        st.stop()

def leer_excel_con_encabezados(archivo):
    """Lee archivo Excel con encabezados detectando el engine correcto"""
    nombre = archivo.name.lower()
    
    try:
        if nombre.endswith('.xls') and not nombre.endswith('.xlsx'):
            try:
                import xlrd
                return pd.read_excel(archivo, sheet_name=0, header=0, engine='xlrd')
            except ImportError:
                st.error("❌ Para archivos .xls es necesario instalar xlrd. Ejecuta: pip install xlrd")
                st.stop()
        else:
            return pd.read_excel(archivo, sheet_name=0, header=0, engine='openpyxl')
    except Exception as e:
        try:
            return pd.read_excel(archivo, sheet_name=0, header=None, engine='openpyxl')
        except:
            st.error(f"No se pudo leer el archivo. Error: {str(e)}")
            st.stop()

# =========================================================
# 🔥 DETECCIÓN DE BANCO POR CONTENIDO DEL ARCHIVO
# =========================================================

def mono_detectar_banco_por_contenido(archivo):
    """
    Detecta el banco leyendo el contenido del archivo, no solo el nombre.
    Soporta formatos binarios de Excel, HTML o texto de forma robusta.
    """
    try:
        # Guardar la posición actual
        pos = archivo.tell()
        archivo.seek(0)
        
        # Leer usando la función robusta mono_leer_excel_sin_encabezados
        df_temp = mono_leer_excel_sin_encabezados(archivo)
        
        if df_temp is not None and not df_temp.empty:
            # Incluir encabezados/columnas de forma robusta (soporta MultiIndex)
            columnas_texto = ""
            if isinstance(df_temp.columns, pd.MultiIndex):
                for lvl in df_temp.columns.levels:
                    columnas_texto += " " + " ".join([str(x) for x in lvl if pd.notna(x)])
            else:
                columnas_texto = " ".join([str(x) for x in df_temp.columns if pd.notna(x)])
            
            # Convertir las primeras 40 filas a string para buscar
            df_sub = df_temp.head(40)
            texto_valores = " ".join([str(val) for val in df_sub.values.flatten() if pd.notna(val)])
            texto = (columnas_texto + " " + texto_valores).upper()
            
            # Restablecer la posición del archivo
            archivo.seek(pos)
            
            # Detectar por contenido
            if "BANCAMIGA" in texto or "BANCAMIGA BANCO UNIVERSAL" in texto or "BANCA AMIGA" in texto or "AMIGA" in texto:
                return "bancamiga"
            elif "BANESCO" in texto:
                return "banesco"
            elif "PROVINCIAL" in texto or "BBVA" in texto or "OPERACIÓN" in texto or "F. VALOR" in texto or "F.OPERACIÓN" in texto:
                return "provincial"
            elif "BANCO DE VENEZUELA" in texto or "BDV" in texto:
                return "venezuela"
            elif "BNC" in texto or "BANCO NACIONAL DE CREDITO" in texto:
                return "bnc"
            elif "BANPLUS" in texto or "BAN PLUS" in texto:
                return "banplus"
            elif "MERCANTIL" in texto:
                return "mercantil"
            elif "TESORO" in texto or "BANCO DEL TESORO" in texto:
                return "tesoro"
                
        # Restablecer si no se detectó nada
        archivo.seek(pos)
        return None
        
    except Exception as e:
        try:
            archivo.seek(pos)
        except Exception:
            pass
        return None

def mono_detectar_banco_por_nombre(nombre_archivo):
    """Detecta el banco por el nombre del archivo (fallback)"""
    nombre = nombre_archivo.upper()

    # Detectar por número de cuenta en el nombre (20 dígitos continuos o separados)
    clean_name = re.sub(r'[\s\-_]', '', nombre_archivo)
    match_banco = re.search(r'(0102|0105|0108|0134|0163|0172|0174|0191)\d{16}', clean_name)
    if match_banco:
        codigo = match_banco.group(1)
        if codigo == "0102":
            return "venezuela"
        elif codigo == "0105":
            return "mercantil"
        elif codigo == "0108":
            return "provincial"
        elif codigo == "0134":
            return "banesco"
        elif codigo == "0163":
            return "tesoro"
        elif codigo == "0172":
            return "bancamiga"
        elif codigo == "0174":
            return "banplus"
        elif codigo == "0191":
            return "bnc"

    if "TESORO" in nombre or "TESORERIA" in nombre or "TES" in nombre:
        return "tesoro"
    elif "BANCAMIGA" in nombre or "BANCAAMIGA" in nombre or "AMIGA" in nombre:
        return "bancamiga"
    elif "BANPLUS" in nombre:
        return "banplus"
    elif "BANESCO" in nombre or re.match(r"^J\d+", nombre_archivo):
        return "banesco"
    elif (
        "MOVIMIENTOS EN MONEDA NACIONAL" in nombre
        or "VENEZUELA" in nombre
        or "BANCO DE VENEZUELA" in nombre
        or "BDV" in nombre
        or "VZLA" in nombre
    ):
        return "venezuela"
    elif "PROVINCIAL" in nombre or "PROV" in nombre:
        return "provincial"
    elif "BNC" in nombre:
        return "bnc"
    elif "MERCANTIL" in nombre:
        return "mercantil"
    return "mercantil"

# =========================================================
# FUNCIONES ORIGINALES (NO MODIFICADAS)
# =========================================================

def mono_convertir_monto(valor):
    try:
        if pd.isna(valor):
            return None

        if isinstance(valor, (int, float)):
            numero = float(valor)
            if isinstance(valor, int) and numero >= 100000:
                numero = numero / 100
            return numero

        valor_original = str(valor).strip()
        valor = valor_original
        valor = valor.replace(" ", "")
        valor = valor.replace("$", "")
        valor = valor.replace("Bs", "")
        valor = valor.replace("€", "")

        if valor == "":
            return None

        if "." in valor and "," in valor:
            valor = valor.replace(".", "")
            valor = valor.replace(",", ".")
        elif "," in valor:
            valor = valor.replace(",", ".")

        numero = float(valor)

        if "." not in valor_original and "," not in valor_original and numero >= 100000:
            numero = numero / 100

        return numero

    except Exception:
        return None

def mono_limpiar_monto_banplus(valor):
    if valor is None or pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    
    valor_str = str(valor).strip()
    if not valor_str:
        return 0.0
    
    valor_str = valor_str.replace('$', '').replace('Bs.', '').replace('Bs', '').replace(' ', '').strip()
    
    try:
        return float(valor_str)
    except ValueError:
        pass
    
    has_comma = ',' in valor_str
    has_dot = '.' in valor_str
    
    if has_comma and has_dot:
        pos_comma = valor_str.rfind(',')
        pos_dot = valor_str.rfind('.')
        if pos_dot > pos_comma:
            valor_limpio = valor_str.replace(',', '')
        else:
            valor_limpio = valor_str.replace('.', '').replace(',', '.')
    elif has_comma:
        valor_limpio = valor_str.replace(',', '.')
    elif has_dot:
        valor_limpio = valor_str
    else:
        valor_limpio = valor_str
        
    try:
        return float(valor_limpio)
    except ValueError:
        return 0.0

# =========================================================
# CALCULAR USD SEGÚN TASA
# =========================================================

def calcular_usd(monto_bs, tasa):
    try:
        if monto_bs is None or tasa is None or tasa == 0:
            return None
        return round(abs(monto_bs) / abs(tasa), 2)
    except:
        return None

# =========================================================
# 🔥 DETECTAR COMISIONES - VERSIÓN MEJORADA CON PALABRAS CLAVE
# =========================================================

def mono_es_comision(texto, proveedor=None):
    """
    Detecta si un movimiento es una comisión bancaria.
    
    REGLAS:
    - Si tiene proveedor asociado → NO es comisión bancaria (es un pago a terceros)
    - Si es pago a personal (nómina, comisiones de ventas) → NO es comisión bancaria
    - Solo son comisiones bancarias: cargos del banco (ITF, mantenimiento, comisión por transferencia, etc.)
    """
    texto = normalizar_texto(texto).strip()
    texto_upper = texto.upper()
    
    # 🔥 REGLA 1: Si tiene proveedor asociado, NO es comisión bancaria
    if proveedor and str(proveedor).strip():
        return False
    
    # 🔥 REGLA 2: COMISIONES PAGADAS A PERSONAL = EGRESO (no comisión bancaria)
    if any(x in texto for x in [
        "comisiones sobre servicios contratados",
        "comision gerente comercial",
        "comision vendedor",
        "comision asesor",
        "comision ejecutivo",
        "comision supervisor",
        "comision ventas",
        "comisiones ventas",
        "comision comercial",
        "comisiones comerciales"
    ]):
        return False
    
    # 🔥 REGLA 3: NUNCA SON COMISIONES BANCARIAS
    if any(x in texto for x in [
        "pago a proveedores",
        "pago de nomina",
        "nomina",
        "transf entre ctas",
        "transferencia a terceros",
        "pago movil comercial"
    ]):
        return False
    
    # 🔥 REGLA 4: SOLO SON COMISIONES BANCARIAS si coinciden con estas palabras
    palabras_comision_bancaria = [
        "comision por transferencia",
        "comision pago movil",
        "comisión pago movil",
        "comision x pago de nomina",
        "comision x pago de nominas",
        "itf",
        "impuesto a las transacciones financieras",
        "cargo bancario",
        "mantenimiento de cuenta",
        "comision bancaria",
        "comisión bancaria",
        "cargo por servicio",
        "cargo por transaccion",
        "comision por",
        "comisión por",
        "comision pago movil comercial",
        "comision x pago de nominas mb",
        "com pago otras ctas",
        "com pago otr bcos",
        "comis",
        "comis. cr.i",
        "sms",
        "servicio sms",
        "servicio sms plus",
        "sms plus",
        "domiciliacion j412438905",
        "distribuidora global",
        "emision edo",
        "retencion de impuesto",
        "com. trf",
        "com.serv",
        "emision de estado",
        "below minimum balance charges",
        "stament service",
        "statement service"
    ]
    
    # Verificar si coincide con alguna comisión bancaria
    for patron in palabras_comision_bancaria:
        if patron in texto:
            return True
    
    # Si contiene "comision" pero no coincide con las reglas anteriores, NO es comisión bancaria
    if "comision" in texto or "comisión" in texto:
        return False
    
    return False

# =========================================================
# 🔥 FUNCIÓN MEJORADA: ENRIQUECER EGRESOS CON IPAGO (CRUCE FLEXIBLE)
# =========================================================

def mono_enriquecer_egresos_con_ipago(df_egresos, df_ipago):
    """
    Enriquece los egresos del banco con datos de iPago.
    Usa múltiples estrategias de cruce:
    1. Coincidencia exacta de referencia
    2. Coincidencia parcial (quitando ceros o X)
    3. Coincidencia por monto + fecha
    """
    if df_ipago is None or df_ipago.empty:
        return df_egresos
    
    # Hacer una copia para no modificar el original
    df_resultado = df_egresos.copy()
    
    # 🔥 NORMALIZAR REFERENCIAS DEL BANCO
    df_resultado["REFERENCIA_NORM"] = (
        df_resultado["REFERENCIA"]
        .astype(str)
        .str.replace(".0", "", regex=False)
        .str.strip()
    )
    
    # 🔥 NORMALIZAR REFERENCIAS DE IPAGO
    df_ipago["Referencia_Norm"] = (
        df_ipago["Referencia"]
        .astype(str)
        .str.replace(".0", "", regex=False)
        .str.strip()
    )
    
    # 🔥 CREAR VARIANTES DE REFERENCIA PARA CRUCE FLEXIBLE
    def generar_variantes_ref(ref):
        ref = str(ref).strip()
        if not ref or ref == "nan":
            return set()
        
        variantes = set()
        variantes.add(ref)  # Original
        
        # Quitar ceros a la izquierda
        ref_sin_ceros = ref.lstrip('0')
        if ref_sin_ceros != ref and ref_sin_ceros:
            variantes.add(ref_sin_ceros)
        
        # Quitar X al final
        if ref.endswith('X'):
            ref_sin_x = ref[:-1]
            if ref_sin_x:
                variantes.add(ref_sin_x)
                # También sin ceros
                ref_sin_x_sin_ceros = ref_sin_x.lstrip('0')
                if ref_sin_x_sin_ceros:
                    variantes.add(ref_sin_x_sin_ceros)
        
        # Si tiene X pero no es el final (caso raro)
        if 'X' in ref and not ref.endswith('X'):
            ref_sin_x = ref.replace('X', '')
            if ref_sin_x:
                variantes.add(ref_sin_x)
                ref_sin_x_sin_ceros = ref_sin_x.lstrip('0')
                if ref_sin_x_sin_ceros:
                    variantes.add(ref_sin_x_sin_ceros)
        
        # 🔥 NUEVO PARA VENEZUELA: Si la referencia tiene 11 dígitos y comienza con 0
        if len(ref) >= 10 and ref.startswith('0'):
            ref_sin_cero_inicial = ref[1:]
            if ref_sin_cero_inicial:
                variantes.add(ref_sin_cero_inicial)
                variantes.add(ref_sin_cero_inicial.lstrip('0'))
        
        # 🔥 NUEVO PARA VENEZUELA: Si la referencia tiene formato numérico con puntos
        if '.' in ref:
            ref_sin_puntos = ref.replace('.', '')
            if ref_sin_puntos:
                variantes.add(ref_sin_puntos)
                variantes.add(ref_sin_puntos.lstrip('0'))
        
        return variantes
    
    # 🔥 CREAR DICCIONARIO DE IPAGO CON TODAS LAS VARIANTES
    ipago_dict = {}
    for _, row in df_ipago.iterrows():
        ref_original = str(row.get("Referencia", "")).strip()
        
        # Si la referencia es NaN o está vacía, usar monto como clave
        if not ref_original or ref_original == "nan":
            continue
        
        # Generar todas las variantes de esta referencia
        variantes = generar_variantes_ref(ref_original)
        
        for variante in variantes:
            if variante and variante not in ipago_dict:
                ipago_dict[variante] = {
                    "PROVEEDOR": row.get("Proveedor", ""),
                    "TIPO_EGRESO": row.get("Tipo de Egreso", ""),
                    "TIPO_PAGO": row.get("Tipo de Pago", ""),
                    "DESCRIPCION_IPAGO": row.get("Descripción", ""),
                    "FECHA_PAGO": row.get("Fecha Pago", ""),
                    "EMPRESA": row.get("Empresa", ""),
                    "MONTO_IPAGO": row.get("Monto", 0),
                    "MONTO_USD": row.get("Monto USD", 0),
                    "REFERENCIA_ORIGINAL": ref_original
                }
    
    # 🔥 ENRIQUECER CADA EGRESO
    for idx, row in df_resultado.iterrows():
        ref_banco = str(row.get("REFERENCIA", "")).strip()
        monto_banco = float(row.get("MONTO BS", 0))
        fecha_banco = str(row.get("FECHA", ""))
        
        # GENERAR VARIANTES DE LA REFERENCIA DEL BANCO
        variantes_banco = generar_variantes_ref(ref_banco)
        
        # BUSCAR COINCIDENCIA POR REFERENCIA
        coincide_ref = False
        datos_encontrados = None
        
        for variante in variantes_banco:
            if variante in ipago_dict:
                datos_encontrados = ipago_dict[variante]
                coincide_ref = True
                break
        
        # SI NO COINCIDE POR REFERENCIA, INTENTAR POR MONTO + FECHA
        if not coincide_ref:
            # Buscar en iPago por monto similar (con margen de 1%)
            for clave, datos in ipago_dict.items():
                monto_ipago = float(datos.get("MONTO_IPAGO", 0))
                if monto_ipago > 0:
                    diferencia = abs(monto_banco - monto_ipago) / max(monto_banco, monto_ipago)
                    if diferencia < 0.01:  # 1% de margen
                        datos_encontrados = datos
                        coincide_ref = True
                        break
        
        # APLICAR DATOS ENCONTRADOS
        if datos_encontrados:
            df_resultado.at[idx, "STATUS"] = datos_encontrados["PROVEEDOR"]
            df_resultado.at[idx, "OBSERVACIÓN"] = datos_encontrados["TIPO_EGRESO"]
            df_resultado.at[idx, "TIPO_PAGO"] = datos_encontrados["TIPO_PAGO"]
            df_resultado.at[idx, "PROVEEDOR_IPAGO"] = datos_encontrados["PROVEEDOR"]
            df_resultado.at[idx, "REFERENCIA_IPAGO"] = datos_encontrados.get("REFERENCIA_ORIGINAL", "")
            
            # 🔥 Reemplazar descripción con la de iPago
            descripcion_ipago = datos_encontrados["DESCRIPCION_IPAGO"]
            if descripcion_ipago:
                df_resultado.at[idx, "DESCRIPCIÓN"] = descripcion_ipago
                df_resultado.at[idx, "DESCRIPCION_ORIGINAL"] = row.get("DESCRIPCIÓN", "")
            
            # 🔥 Si es comisión, marcarlo como tal
            tipo = str(datos_encontrados["TIPO_EGRESO"]).upper()
            desc = str(datos_encontrados["DESCRIPCION_IPAGO"]).upper()
            if "COMISION" in tipo or "COMISION" in desc:
                df_resultado.at[idx, "ES_COMISION"] = True
            else:
                df_resultado.at[idx, "ES_COMISION"] = False
        else:
            # Si no hay coincidencia, mantener los valores actuales
            df_resultado.at[idx, "STATUS"] = "SIN DATOS IPAGO"
            df_resultado.at[idx, "OBSERVACIÓN"] = "SIN CONCORDANCIA"
            df_resultado.at[idx, "TIPO_PAGO"] = ""
            df_resultado.at[idx, "PROVEEDOR_IPAGO"] = ""
            df_resultado.at[idx, "ES_COMISION"] = False
    
    # Eliminar columnas auxiliares
    df_resultado = df_resultado.drop(
        columns=["REFERENCIA_NORM"], 
        errors="ignore"
    )
    
    return df_resultado

# =========================================================
# PROCESAR BANESCO
# =========================================================

def mono_procesar_banesco(df):
    st.info("Procesando Banesco...")
    try:
        df.columns = ["FECHA", "REFERENCIA", "DESCRIPCION", "MONTO_RAW", "BALANCE"]
        df.columns = [str(c).strip().upper() for c in df.columns]

        rename_map = {}
        for col in df.columns:
            c = str(col).lower()
            if "fecha" in c:
                rename_map[col] = "FECHA"
            elif "referencia" in c:
                rename_map[col] = "REFERENCIA"
            elif "descrip" in c:
                rename_map[col] = "DESCRIPCION"
            elif "monto" in c:
                rename_map[col] = "MONTO_RAW"

        df = df.rename(columns=rename_map)

        for col in ["FECHA", "REFERENCIA", "DESCRIPCION", "MONTO_RAW"]:
            if col not in df.columns:
                st.error(f"No existe columna: {col}")
                return pd.DataFrame()

        # Convertir fechas de manera robusta
        def parse_banesco_date(val):
            val_str = str(val).strip()
            if not val_str or val_str == "nan":
                return pd.NaT
            if len(val_str) >= 5 and val_str[:4].isdigit() and val_str[4] in ('/', '-'):
                return pd.to_datetime(val_str, dayfirst=False, errors="coerce")
            return pd.to_datetime(val_str, dayfirst=True, errors="coerce")

        df["FECHA"] = df["FECHA"].apply(parse_banesco_date)
        df = df[df["FECHA"].notna()]

        df["TIPO"] = df["MONTO_RAW"].astype(str).apply(lambda x: "NC" if "+" in x else "ND")
        df["MONTO"] = (df["MONTO_RAW"].astype(str).str.replace("+", "", regex=False)
                       .str.replace("-", "", regex=False)
                       .str.replace(".", "", regex=False)
                       .str.replace(",", ".", regex=False)
                       .str.strip())

        df["MONTO"] = pd.to_numeric(df["MONTO"], errors="coerce")
        df = df[df["MONTO"].notna()]
        df = df[df["MONTO"] > 0]

        df = df[["FECHA", "REFERENCIA", "DESCRIPCION", "TIPO", "MONTO"]]
        st.success(f"Banesco OK: {len(df)} movimientos")
        st.dataframe(df.head())
        return df

    except Exception as e:
        st.error(f"Error Banesco: {str(e)}")
        return pd.DataFrame()

# =========================================================
# PROCESAR PROVINCIAL - VERSIÓN MEJORADA PARA FORMATO ESPECÍFICO
# =========================================================

def mono_procesar_provincial(df):
    """
    Procesa archivo de Provincial con formato específico.
    El archivo tiene un formato de texto con columnas:
    F. Operación | F. Valor | Código | Nº. Doc. | Concepto | Importe | Oficina
    """
    st.info("🔍 Procesando archivo de Provincial (formato especial)...")
    
    try:
        # Mostrar información del archivo
        st.write("📊 **Información del archivo:**")
        st.write(f"- Número de filas: {len(df)}")
        st.write(f"- Número de columnas: {len(df.columns)}")
        
        # Mostrar primeras filas para debug
        st.write("👁️ **Primeras 15 filas del archivo:**")
        st.dataframe(df.head(15))
        
        # Buscar la fila que contiene los encabezados
        encabezado_idx = None
        for i in range(min(30, len(df))):
            fila = df.iloc[i]
            # Convertir TODOS los valores a string para evitar errores
            fila_str = [str(val) for val in fila.tolist()]
            texto_fila = " ".join(fila_str).upper()
            
            # Buscar columnas que contengan "F. Operación" o "Concepto" o "Importe"
            if "CONCEPTO" in texto_fila and "IMPORTE" in texto_fila:
                encabezado_idx = i
                break
        
        if encabezado_idx is None:
            st.error("❌ No se encontró la fila de encabezados en el archivo Provincial.")
            return pd.DataFrame()
        
        st.write(f"✅ Encabezados encontrados en la fila {encabezado_idx}")
        
        # Obtener los encabezados
        headers = df.iloc[encabezado_idx].astype(str).str.strip().tolist()
        st.write("📋 **Encabezados detectados:**", headers)
        
        # Limpiar y mapear encabezados
        rename_map = {}
        for col in headers:
            col_clean = str(col).strip().upper()
            if "OPERAC" in col_clean or "FECHA" in col_clean:
                rename_map[col] = "FECHA"
            elif "F. VALOR" in col_clean:
                rename_map[col] = "FECHA_VALOR"
            elif "CÓDIGO" in col_clean or "CODIGO" in col_clean:
                rename_map[col] = "CODIGO"
            elif "Nº. DOC" in col_clean or "NRO DOC" in col_clean or "DOC" in col_clean:
                rename_map[col] = "REFERENCIA"
            elif "CONCEPTO" in col_clean:
                rename_map[col] = "DESCRIPCION"
            elif "IMPORTE" in col_clean:
                rename_map[col] = "MONTO"
            elif "OFICINA" in col_clean:
                rename_map[col] = "OFICINA"
        
        st.write("📋 **Mapeo de columnas:**", rename_map)
        
        # Asignar encabezados al DataFrame
        df.columns = headers
        df = df.iloc[encabezado_idx + 1:].reset_index(drop=True)
        
        # Renombrar columnas
        df = df.rename(columns=rename_map)
        
        # Verificar columnas necesarias
        if "FECHA" not in df.columns:
            # Intentar encontrar fecha en otra columna
            for col in df.columns:
                if "FECHA" in str(col).upper():
                    df = df.rename(columns={col: "FECHA"})
                    break
        
        if "FECHA" in df.columns:
            # Procesar fechas - convertir a string primero
            df["FECHA"] = df["FECHA"].astype(str).str.strip()
            # Eliminar filas con fechas vacías o que sean encabezados
            df = df[~df["FECHA"].str.contains("FECHA|SALDO|Período", case=False, na=False)]
            # Eliminar filas con fechas que sean números o NaN
            df = df[df["FECHA"].str.match(r'^\d{2}[-/]\d{2}[-/]\d{2,4}$', na=False)]
            
            # Convertir fechas (formato DD-MM-YYYY o DD/MM/YYYY)
            df["FECHA"] = pd.to_datetime(df["FECHA"], dayfirst=True, errors="coerce")
            df = df[df["FECHA"].notna()]
        else:
            st.error("❌ No se encontró columna FECHA en el archivo Provincial.")
            return pd.DataFrame()
        
        # Procesar el monto
        if "MONTO" in df.columns:
            # Limpiar el monto (quitar espacios, puntos, comas) - convertir a string primero
            df["MONTO"] = df["MONTO"].astype(str).str.replace(" ", "", regex=False)
            df["MONTO"] = df["MONTO"].str.replace(".", "", regex=False)
            df["MONTO"] = df["MONTO"].str.replace(",", ".", regex=False)
            df["MONTO"] = df["MONTO"].str.replace("'", "", regex=False)
            
            # Convertir directamente a numérico (sin filtro regex)
            df["MONTO"] = pd.to_numeric(df["MONTO"], errors="coerce")
            
            # Eliminar filas con monto NaN
            df = df[df["MONTO"].notna()]
            
            # Si el monto es negativo, es un ND (débito), si es positivo es NC (crédito)
            df["TIPO"] = df["MONTO"].apply(lambda x: "NC" if x > 0 else "ND" if x < 0 else "")
            
            # Tomar valor absoluto
            df["MONTO"] = df["MONTO"].abs()
            
            # Eliminar filas con monto 0
            df = df[df["MONTO"] > 0]
        else:
            st.error("❌ No se encontró columna MONTO en el archivo Provincial.")
            return pd.DataFrame()
        
        # Asegurar que existe columna REFERENCIA
        if "REFERENCIA" not in df.columns:
            df["REFERENCIA"] = ""
        else:
            df["REFERENCIA"] = df["REFERENCIA"].astype(str).str.strip()
            # Limpiar referencias (quitar comillas simples)
            df["REFERENCIA"] = df["REFERENCIA"].str.replace("'", "", regex=False)
        
        # Asegurar que existe columna DESCRIPCION
        if "DESCRIPCION" not in df.columns:
            df["DESCRIPCION"] = ""
        else:
            df["DESCRIPCION"] = df["DESCRIPCION"].astype(str).str.strip()
        
        # 🔥 DETECTAR COMISIONES DE PROVINCIAL
        df["ES_COMISION"] = df["DESCRIPCION"].str.contains("COMIS", case=False, na=False)
        
        # 🔥 DEBUG: Mostrar cuántas comisiones se detectaron
        num_comisiones = df["ES_COMISION"].sum()
        st.info(f"💳 Se detectaron {num_comisiones} comisiones en el archivo Provincial")
        
        # Mostrar las comisiones detectadas
        if num_comisiones > 0:
            st.write("📋 **Comisiones detectadas:**")
            st.dataframe(df[df["ES_COMISION"] == True][["FECHA", "REFERENCIA", "DESCRIPCION", "MONTO"]])
        
        # Seleccionar solo las columnas necesarias
        df_resultado = df[["FECHA", "REFERENCIA", "DESCRIPCION", "TIPO", "MONTO", "ES_COMISION"]].copy()
        
        # Mostrar resultados
        st.success(f"✅ Provincial OK: {len(df_resultado)} movimientos detectados")
        st.dataframe(df_resultado.head(10))
        
        return df_resultado
        
    except Exception as e:
        st.error(f"❌ Error procesando Provincial: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return pd.DataFrame()

# =========================================================
# PROCESAR BNC
# =========================================================

def mono_procesar_bnc(df):
    st.info("Procesando archivo BNC...")
    encabezado = None

    for i in range(min(30, len(df))):
        fila = df.iloc[i].fillna("").astype(str)
        texto = " ".join(fila.tolist()).lower()
        if "fecha" in texto and ("descripcion" in texto or "descripción" in texto):
            encabezado = i
            break

    if encabezado is None:
        st.error("No se encontró encabezado válido en BNC")
        return pd.DataFrame()

    headers = []
    for idx, col in enumerate(df.iloc[encabezado]):
        col = str(col).strip().replace("\n", " ")
        if col == "" or col.lower() == "nan":
            col = f"COLUMNA_{idx}"
        headers.append(col)

    headers_unicos = []
    contador = {}
    for h in headers:
        if h in contador:
            contador[h] += 1
            nuevo = f"{h}_{contador[h]}"
        else:
            contador[h] = 0
            nuevo = h
        headers_unicos.append(nuevo)

    df.columns = headers_unicos
    df = df.iloc[encabezado + 1:].reset_index(drop=True)

    rename_map = {}
    for col in df.columns:
        col_str = str(col).strip().lower()
        if "fecha" in col_str:
            rename_map[col] = "FECHA"
        elif "descripcion" in col_str or "descripción" in col_str or "concepto" in col_str:
            rename_map[col] = "DESCRIPCION"
        elif "referencia" in col_str:
            rename_map[col] = "REFERENCIA"
        elif "credito" in col_str or "haber" in col_str:
            rename_map[col] = "CREDITO"
        elif "debito" in col_str or "debe" in col_str:
            rename_map[col] = "DEBITO"

    df = df.rename(columns=rename_map)

    if "FECHA" in df.columns:
        df["FECHA"] = pd.to_datetime(df["FECHA"], dayfirst=True, errors="coerce")
        df = df[df["FECHA"].notna()]

    df["CREDITO"] = pd.to_numeric(df.get("CREDITO", 0), errors="coerce").fillna(0) if "CREDITO" in df.columns else 0
    df["DEBITO"] = pd.to_numeric(df.get("DEBITO", 0), errors="coerce").fillna(0) if "DEBITO" in df.columns else 0

    df["MONTO"] = df["CREDITO"] - df["DEBITO"]
    df["TIPO"] = df["MONTO"].apply(lambda x: "NC" if x > 0 else "ND")
    df["MONTO"] = df["MONTO"].abs()
    df = df[df["MONTO"] != 0]

    st.success(f"Registros BNC: {len(df)}")
    st.dataframe(df.head())
    return df

# =========================================================
# PROCESAR TESORO
# =========================================================

def mono_procesar_tesoro(df):
    st.info("Procesando Banco del Tesoro...")
    try:
        encabezado = None
        for i in range(min(20, len(df))):
            fila = df.iloc[i].astype(str)
            texto = " ".join(map(str, fila.tolist())).lower()
            if "fecha" in texto and "referencia" in texto and "concepto" in texto:
                encabezado = i
                break

        if encabezado is None:
            st.error("No se encontró encabezado válido en Tesoro")
            return pd.DataFrame()

        df.columns = df.iloc[encabezado]
        df = df.iloc[encabezado + 1:].reset_index(drop=True)
        df.columns = [str(c).strip() for c in df.columns]

        rename_map = {}
        for col in df.columns:
            c = str(col).strip().lower()
            if "fecha" in c:
                rename_map[col] = "FECHA"
            elif "referencia" in c:
                rename_map[col] = "REFERENCIA"
            elif "concepto" in c:
                rename_map[col] = "DESCRIPCION"
            elif "débito" in c or "debito" in c:
                rename_map[col] = "DEBITO"
            elif "crédito" in c or "credito" in c:
                rename_map[col] = "CREDITO"
            elif "código" in c or "codigo" in c:
                rename_map[col] = "TIPO"

        df = df.rename(columns=rename_map)

        if "FECHA" not in df.columns:
            st.error("No existe columna FECHA")
            return pd.DataFrame()

        df["FECHA"] = pd.to_datetime(df["FECHA"], dayfirst=True, errors="coerce")
        df = df[df["FECHA"].notna()]

        def limpiar_numero(valor):
            valor = str(valor).replace(".", "").replace(",", ".")
            try:
                return float(valor)
            except:
                return 0

        df["CREDITO"] = df.get("CREDITO", 0).apply(limpiar_numero) if "CREDITO" in df.columns else 0
        df["DEBITO"] = df.get("DEBITO", 0).apply(limpiar_numero) if "DEBITO" in df.columns else 0

        df["MONTO"] = df["CREDITO"] - df["DEBITO"]
        df["TIPO"] = df["MONTO"].apply(lambda x: "NC" if x > 0 else "ND")
        df["MONTO"] = df["MONTO"].abs()
        df = df[df["MONTO"] > 0]

        df = df[["FECHA", "REFERENCIA", "DESCRIPCION", "TIPO", "MONTO"]]
        st.success(f"Tesoro OK: {len(df)} registros")
        st.dataframe(df.head())
        return df

    except Exception as e:
        st.error(f"Error Tesoro: {str(e)}")
        return pd.DataFrame()

# =========================================================
# PROCESAR BANCAMIGA - VERSIÓN MEJORADA CON DETECCIÓN DE ENCABEZADOS
# =========================================================

def mono_procesar_bancamiga(df):
    """
    Procesa archivo de Bancamiga con formato específico.
    El archivo tiene columnas: Nro. | Fecha | Referencia | Concepto | Débito | Crédito | Saldo
    """
    st.info("🔍 Procesando archivo de Bancamiga...")
    
    try:
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(-1)
        # Mostrar información del archivo
        st.write("📊 **Información del archivo:**")
        st.write(f"- Número de filas: {len(df)}")
        st.write(f"- Número de columnas: {len(df.columns)}")
        
        # Mostrar primeras filas para debug
        st.write("👁️ **Primeras 15 filas del archivo:**")
        st.dataframe(df.head(15))
        
        # 🔥 VERIFICAR SI YA TIENE ENCABEZADOS
        columnas = [str(c).strip().upper() for c in df.columns]
        
        if "FECHA" in columnas and "REFERENCIA" in columnas:
            st.success("✅ Encabezados ya presentes.")
            
            # Mapeo de columnas
            rename_map = {
                "NRO.": "NRO",
                "NRO": "NRO",
                "FECHA": "FECHA",
                "REFERENCIA": "REFERENCIA",
                "CONCEPTO": "DESCRIPCION",
                "DÉBITO": "DEBITO",
                "DEBITO": "DEBITO",
                "CRÉDITO": "CREDITO",
                "CREDITO": "CREDITO",
                "SALDO": "SALDO"
            }
            
            df.columns = [rename_map.get(str(c).strip().upper(), str(c).strip().upper())
                          for c in df.columns]
            
        else:
            st.info("🔍 Buscando fila de encabezados...")
            
            # Buscar la fila que contiene los encabezados
            encabezado_idx = None
            for i in range(min(30, len(df))):
                fila = df.iloc[i]
                # Convertir TODOS los valores a string para evitar errores
                fila_str = [str(val) for val in fila.tolist()]
                texto_fila = " ".join(fila_str).upper()
                
                # Buscar columnas que contengan "NRO" o "FECHA" o "REFERENCIA" o "CONCEPTO"
                if "NRO" in texto_fila and "FECHA" in texto_fila and "REFERENCIA" in texto_fila:
                    encabezado_idx = i
                    break
            
            if encabezado_idx is None:
                st.error("❌ No se encontró la fila de encabezados en el archivo Bancamiga.")
                return pd.DataFrame()
            
            st.write(f"✅ Encabezados encontrados en la fila {encabezado_idx}")
            
            # Obtener los encabezados
            headers = df.iloc[encabezado_idx].astype(str).str.strip().tolist()
            st.write("📋 **Encabezados detectados:**", headers)
            
            # Limpiar y mapear encabezados
            rename_map = {}
            for col in headers:
                col_clean = str(col).strip().upper()
                if "NRO" in col_clean or "Nº" in col_clean:
                    rename_map[col] = "NRO"
                elif "FECHA" in col_clean:
                    rename_map[col] = "FECHA"
                elif "REFERENCIA" in col_clean:
                    rename_map[col] = "REFERENCIA"
                elif "CONCEPTO" in col_clean:
                    rename_map[col] = "DESCRIPCION"
                elif "DÉBITO" in col_clean or "DEBITO" in col_clean:
                    rename_map[col] = "DEBITO"
                elif "CRÉDITO" in col_clean or "CREDITO" in col_clean:
                    rename_map[col] = "CREDITO"
                elif "SALDO" in col_clean:
                    rename_map[col] = "SALDO"
            
            st.write("📋 **Mapeo de columnas:**", rename_map)
            
            # Asignar encabezados al DataFrame
            df.columns = headers
            df = df.iloc[encabezado_idx + 1:].reset_index(drop=True)
            
            # Renombrar columnas
            df = df.rename(columns=rename_map)
        
        # Verificar columnas necesarias
        if "FECHA" not in df.columns:
            st.error("❌ No se encontró columna FECHA en el archivo Bancamiga.")
            return pd.DataFrame()
        
        # 🔥 PROCESAR FECHAS DE BANCAMIGA DE MANERA ROBUSTA
        # Filtrar filas que no son movimientos
        fechas_str_col = df["FECHA"].astype(str).str.strip()
        df = df[
            ~fechas_str_col.str.contains(
                "FECHA|SALDO|TOTAL|CRÉDITO|CREDITO|DÉBITO|DEBITO",
                case=False,
                na=False
            )
        ]

        # Convertir fechas de manera robusta
        df["FECHA_DT"] = pd.to_datetime(df["FECHA"], dayfirst=True, errors="coerce")
        mask = df["FECHA_DT"].isna()
        if mask.any():
            df.loc[mask, "FECHA_DT"] = pd.to_datetime(
                df.loc[mask, "FECHA"].astype(str).str.strip(),
                dayfirst=True,
                errors="coerce"
            )
        
        df = df[df["FECHA"].notna()]
        
        # Procesar débito y crédito
        def limpiar_monto(val):
            val_str = str(val).strip().replace(" ", "")
            if not val_str or val_str == "nan":
                return 0.0
            if "," in val_str:
                val_str = val_str.replace(".", "").replace(",", ".")
            return pd.to_numeric(val_str, errors="coerce")

        if "DEBITO" in df.columns:
            df["DEBITO"] = df["DEBITO"].apply(limpiar_monto).fillna(0)
        else:
            df["DEBITO"] = 0.0
        
        if "CREDITO" in df.columns:
            df["CREDITO"] = df["CREDITO"].apply(limpiar_monto).fillna(0)
        else:
            df["CREDITO"] = 0.0
        
        # Determinar tipo y monto
        df["MONTO"] = df["CREDITO"] - df["DEBITO"]
        df["TIPO"] = df["MONTO"].apply(lambda x: "NC" if x > 0 else "ND" if x < 0 else "")
        df["MONTO"] = df["MONTO"].abs()
        
        # Eliminar filas con monto 0
        df = df[df["MONTO"] > 0]
        
        # Asegurar que existe columna REFERENCIA
        if "REFERENCIA" not in df.columns:
            df["REFERENCIA"] = ""
        else:
            df["REFERENCIA"] = df["REFERENCIA"].astype(str).str.strip()
            # Limpiar referencias (quitar comillas simples)
            df["REFERENCIA"] = df["REFERENCIA"].str.replace("'", "", regex=False)
        
        # Asegurar que existe columna DESCRIPCION
        if "DESCRIPCION" not in df.columns:
            df["DESCRIPCION"] = ""
        else:
            df["DESCRIPCION"] = df["DESCRIPCION"].astype(str).str.strip()
        
        # 🔥 DETECTAR COMISIONES DE BANCAMIGA
        # Las comisiones tienen "Comisión" en la descripción
        df["ES_COMISION"] = df["DESCRIPCION"].str.contains("Comisi", case=False, na=False)
        
        # 🔥 DEBUG: Mostrar cuántas comisiones se detectaron
        num_comisiones = df["ES_COMISION"].sum()
        st.info(f"💳 Se detectaron {num_comisiones} comisiones en el archivo Bancamiga")
        
        # Mostrar las comisiones detectadas
        if num_comisiones > 0:
            st.write("📋 **Comisiones detectadas:**")
            st.dataframe(df[df["ES_COMISION"] == True][["FECHA", "REFERENCIA", "DESCRIPCION", "MONTO"]])
        
        # Seleccionar solo las columnas necesarias
        df_resultado = df[["FECHA", "REFERENCIA", "DESCRIPCION", "TIPO", "MONTO", "ES_COMISION"]].copy()
        
        # Mostrar resultados
        st.success(f"✅ Bancamiga OK: {len(df_resultado)} movimientos detectados")
        st.dataframe(df_resultado.head(10))
        
        return df_resultado
        
    except Exception as e:
        st.error(f"❌ Error procesando Bancamiga: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return pd.DataFrame()

# =========================================================
# OBTENER TASA BCV
# =========================================================

@st.cache_data(ttl=3600)
def mono_obtener_tasa_bcv_fecha(fecha_obj):
    tasas_bcv_local = {
        "01/07/2026": 633.3644,
        "02/07/2026": 639.7029,
        "03/07/2026": 652.9726,
        "04/07/2026": 667.0500,
        "05/07/2026": 667.0500,
        "06/07/2026": 667.0500,
        "07/07/2026": 674.9305,
        "08/07/2026": 685.9427,
        "09/07/2026": 700.2249,
        "10/07/2026": 709.6935,
        "11/07/2026": 721.3456,
        "12/07/2026": 721.3456,
        "13/07/2026": 721.3456,
        "14/06/2026": 587.4059,
        "15/06/2026": 587.4059,
        "16/06/2026": 592.5163,
        "17/06/2026": 596.7824,
        "18/06/2026": 602.3324,
        "19/06/2026": 607.3919,
        "20/06/2026": 612.4332,
        "21/06/2026": 612.4332,
        "22/06/2026": 612.4332,
        "23/06/2026": 617.6388,
        "24/06/2026": 621.5299,
        "25/06/2026": 621.5299,
        "26/06/2026": 622.2135,
        "27/06/2026": 623.0223,
        "28/06/2026": 623.0223,
        "29/06/2026": 623.0223,
        "30/06/2026": 623.0223,
    }
    fecha_str = fecha_obj.strftime("%d/%m/%Y")
    return tasas_bcv_local.get(fecha_str, None)

def obtener_tasa_por_fecha(fecha_obj, usar_api=False):
    return mono_obtener_tasa_bcv_fecha(fecha_obj)

# =========================================================
# CONVERTIR A FORMATO MERCANTIL - INCLUYE FLAG DE COMISIONES
# =========================================================

def mono_convertir_a_formato_mercantil(df, banco):
    """Convierte DataFrame de otros bancos al formato que espera procesar_archivo"""
    datos_convertidos = []
    
    for idx, fila in df.iterrows():
        try:
            fecha = fila.get("FECHA", "")
            if pd.isna(fecha):
                continue
            
            if isinstance(fecha, (pd.Timestamp, datetime)):
                fecha_str = fecha.strftime("%d/%m/%Y")
            else:
                fecha_str = str(fecha)
            
            tipo = fila.get("TIPO", "") or ""
            descripcion = fila.get("DESCRIPCION", "") or ""
            referencia = fila.get("REFERENCIA", "") or ""
            monto = fila.get("MONTO", 0) or 0
            
            # 🔥 OBTENER FLAG DE COMISIÓN
            es_comision = fila.get("ES_COMISION", False)
            if isinstance(es_comision, (bool, np.bool_)):
                es_comision = bool(es_comision)
            else:
                es_comision = False
            
            fila_convertida = [
                "",           # col0
                "",           # col1  
                "",           # col2
                fecha_str,    # col3 - FECHA
                referencia,   # col4 - REFERENCIA
                tipo,         # col5 - TIPO (NC/ND)
                descripcion,  # col6 - DESCRIPCION
                monto,        # col7 - MONTO BS
                "",           # col8
                es_comision,  # col9 - ES_COMISION (flag para identificar comisiones)
            ]
            datos_convertidos.append(fila_convertida)
            
        except Exception as e:
            continue
    
    df_convertido = pd.DataFrame(datos_convertidos)
    return df_convertido if len(df_convertido) > 0 else pd.DataFrame()

# =========================================================
# 🔥 PROCESAR VENEZUELA - VERSIÓN MEJORADA (SIN FILTROS EXCESIVOS)
# =========================================================

def mono_procesar_venezuela_simple(df):
    """Procesa el archivo del BDV usando índices fijos (sin encabezados)"""
    st.info("🔍 Procesando Banco de Venezuela (MODO SIMPLE)...")
    
    try:
        # Mostrar información del archivo
        st.write("📊 **Información del archivo:**")
        st.write(f"- Número de filas: {len(df)}")
        st.write(f"- Número de columnas: {len(df.columns)}")
        
        # Mostrar primeras filas
        st.write("👁️ **Primeras 10 filas del archivo (sin encabezados):**")
        st.dataframe(df.head(10))
        
        # Índices fijos para el formato BDV
        col_fecha = 3
        col_ref = 1
        col_desc = 2
        col_tipo = 4
        col_credito = 5
        col_debito = 6
        
        movimientos = []
        filas_procesadas = 0
        
        # Empezar desde la fila 1 (saltar encabezados)
        for idx in range(1, len(df)):
            try:
                fila = df.iloc[idx]
                
                # Verificar que existe fecha
                if pd.isna(fila[col_fecha]):
                    continue
                
                # Obtener fecha
                fecha_raw = str(fila[col_fecha]).strip()
                
                # Intentar parsear fecha
                fecha_val = None
                try:
                    fecha_val = pd.to_datetime(fecha_raw, format="%d/%m/%Y", errors="coerce")
                except:
                    pass
                
                if pd.isna(fecha_val):
                    try:
                        fecha_val = pd.to_datetime(fecha_raw, dayfirst=True, errors="coerce")
                    except:
                        pass
                
                if pd.isna(fecha_val):
                    continue
                
                # Obtener datos
                referencia = str(fila[col_ref]).strip() if pd.notna(fila[col_ref]) else ""
                descripcion = str(fila[col_desc]).strip() if pd.notna(fila[col_desc]) else ""
                tipo_mov = str(fila[col_tipo]).strip().upper() if pd.notna(fila[col_tipo]) else ""
                
                # 🔥 SOLO FILTRAR POR SALDO INICIAL/FINAL EXPLÍCITO
                desc_upper = descripcion.upper()
                if desc_upper in ["SALDO INICIAL", "SALDO FINAL", "TOTALES"]:
                    continue
                
                # Procesar crédito
                val_credito = 0
                if pd.notna(fila[col_credito]):
                    clean_cred = str(fila[col_credito]).strip()
                    clean_cred = clean_cred.replace(".", "").replace(",", ".")
                    try:
                        val_credito = float(clean_cred) if clean_cred else 0
                    except:
                        pass
                
                # Procesar débito
                val_debito = 0
                if pd.notna(fila[col_debito]):
                    clean_deb = str(fila[col_debito]).strip()
                    clean_deb = clean_deb.replace(".", "").replace(",", ".")
                    try:
                        val_debito = float(clean_deb) if clean_deb else 0
                    except:
                        pass
                
                # Determinar tipo y monto
                monto = 0
                tipo = ""
                
                if tipo_mov == "NC":
                    monto = abs(val_credito)
                    tipo = "NC"
                elif tipo_mov == "ND":
                    monto = abs(val_debito)
                    tipo = "ND"
                else:
                    # Si no tiene tipo definido, determinar por crédito/débito
                    if abs(val_credito) > 0:
                        monto = abs(val_credito)
                        tipo = "NC"
                    elif abs(val_debito) > 0:
                        monto = abs(val_debito)
                        tipo = "ND"
                    else:
                        continue
                
                if monto <= 0:
                    continue
                
                filas_procesadas += 1
                
                # Mostrar primeras 5 filas procesadas
                if filas_procesadas <= 5:
                    st.write(f"✅ **Fila {idx} procesada:** Fecha={fecha_val.strftime('%d/%m/%Y')}, Ref={referencia}, Tipo={tipo}, Monto={monto:,.2f}, Desc={descripcion[:50]}")
                
                movimientos.append({
                    "FECHA": fecha_val.strftime("%d/%m/%Y"),
                    "FECHA_OBJ": fecha_val,
                    "REFERENCIA": referencia,
                    "DESCRIPCION": descripcion,
                    "TIPO": tipo,
                    "MONTO": monto
                })
                
            except Exception as e:
                continue
        
        st.write(f"📊 **Filas procesadas exitosamente:** {filas_procesadas}")
        
        df_resultado = pd.DataFrame(movimientos)
        
        if df_resultado.empty:
            st.error("❌ No se encontraron movimientos válidos en el archivo de Venezuela.")
            return pd.DataFrame()
        
        st.success(f"✅ Venezuela OK: {len(df_resultado)} movimientos detectados")
        st.dataframe(df_resultado.head(10))
        return df_resultado
        
    except Exception as e:
        st.error(f"❌ Error general en procesar_venezuela_simple: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return pd.DataFrame()


def mono_convertir_venezuela_a_formato_mercantil(df):
    """Convierte DataFrame de Venezuela al formato Mercantil - SIN AFECTAR A MERCANTIL"""
    datos_convertidos = []
    
    for idx, fila in df.iterrows():
        try:
            # Obtener fecha
            fecha = None
            if "FECHA_OBJ" in fila and pd.notna(fila["FECHA_OBJ"]):
                fecha = fila["FECHA_OBJ"]
            elif "FECHA" in fila and pd.notna(fila["FECHA"]):
                fecha = fila["FECHA"]
            else:
                continue
            
            if pd.isna(fecha):
                continue
            
            # Convertir a string
            if isinstance(fecha, (pd.Timestamp, datetime)):
                fecha_str = fecha.strftime("%d/%m/%Y")
            else:
                try:
                    fecha_dt = pd.to_datetime(fecha, dayfirst=True, errors="coerce")
                    if pd.notna(fecha_dt):
                        fecha_str = fecha_dt.strftime("%d/%m/%Y")
                    else:
                        fecha_str = str(fecha)
                except:
                    fecha_str = str(fecha)
            
            tipo = fila.get("TIPO", "") or ""
            descripcion = fila.get("DESCRIPCION", "") or ""
            referencia = fila.get("REFERENCIA", "") or ""
            monto = fila.get("MONTO", 0) or 0
            
            fila_convertida = [
                "",           # col0
                "",           # col1  
                "",           # col2
                fecha_str,    # col3 - FECHA
                referencia,   # col4 - REFERENCIA
                tipo,         # col5 - TIPO (NC/ND)
                descripcion,  # col6 - DESCRIPCION
                monto,        # col7 - MONTO BS
                "",           # col8
                False,        # col9 - ES_COMISION (siempre False para Venezuela)
            ]
            datos_convertidos.append(fila_convertida)
            
        except Exception as e:
            continue
    
    df_convertido = pd.DataFrame(datos_convertidos)
    return df_convertido if len(df_convertido) > 0 else pd.DataFrame()

def mono_procesar_banplus(df):
    """
    Procesa archivo de Banplus con formato HTML/Excel.
    Columnas: Fecha | Referencia | Descripción | Débito | Crédito | Saldo
    """
    st.info("🔍 Procesando archivo de Banplus...")
    try:
        # Mostrar información del archivo
        st.write("📊 **Información del archivo:**")
        st.write(f"- Número de filas: {len(df)}")
        st.write(f"- Número de columnas: {len(df.columns)}")
        
        # Encontrar la fila de encabezados si no está en las columnas
        encabezado_idx = None
        cols_upper = [str(c).strip().upper() for c in df.columns]
        if "FECHA" in cols_upper and "REFERENCIA" in cols_upper:
            pass
        else:
            for i in range(min(15, len(df))):
                fila = df.iloc[i].astype(str).str.strip().str.upper().tolist()
                fila_str = " ".join(fila)
                if "FECHA" in fila_str and "REFERENCIA" in fila_str:
                    encabezado_idx = i
                    break
            if encabezado_idx is not None:
                df.columns = df.iloc[encabezado_idx].tolist()
                df = df.iloc[encabezado_idx + 1:].reset_index(drop=True)
        
        # Limpiar columnas
        df.columns = [str(c).strip().upper() for c in df.columns]
        
        rename_map = {}
        for col in df.columns:
            col_clean = str(col).strip().upper()
            if "FECHA" in col_clean: rename_map[col] = "FECHA"
            elif "REFERENCIA" in col_clean: rename_map[col] = "REFERENCIA"
            elif "DESCRIP" in col_clean: rename_map[col] = "DESCRIPCION"
            elif "DÉBITO" in col_clean or "DEBITO" in col_clean or "DEB" in col_clean: rename_map[col] = "DEBITO"
            elif "CRÉDITO" in col_clean or "CREDITO" in col_clean or "CRE" in col_clean: rename_map[col] = "CREDITO"
            elif "SALDO" in col_clean: rename_map[col] = "SALDO"
            
        df = df.rename(columns=rename_map)
        
        # Filtrar filas vacías o totales
        if "FECHA" in df.columns:
            df["FECHA"] = df["FECHA"].astype(str).str.strip()
            df = df[~df["FECHA"].str.contains("FECHA|SALDO|Período|Total", case=False, na=False)]
            df = df[df["FECHA"].str.match(r'^\d{2}[-/]\d{2}[-/]\d{2,4}$', na=False)]
            df["FECHA"] = pd.to_datetime(df["FECHA"], dayfirst=True, errors="coerce")
            df = df[df["FECHA"].notna()]
            
        # Reemplazar valores vacíos o nulos en Débito y Crédito
        for col in ["DEBITO", "CREDITO"]:
            if col in df.columns:
                df[col] = df[col].apply(mono_limpiar_monto_banplus).fillna(0.0)
                
        datos_normalizados = []
        for idx, fila in df.iterrows():
            fecha_str = fila["FECHA"].strftime("%d/%m/%Y")
            referencia = str(fila.get("REFERENCIA", "")).strip().replace("'", "")
            descripcion = str(fila.get("DESCRIPCION", "")).strip()
            
            debito = float(fila.get("DEBITO", 0.0))
            credito = float(fila.get("CREDITO", 0.0))
            
            if credito > 0:
                tipo = "NC"
                monto = credito
            elif debito > 0:
                tipo = "ND"
                monto = debito
            else:
                continue
                
            datos_normalizados.append({
                "FECHA": fecha_str,
                "REFERENCIA": referencia,
                "TIPO": tipo,
                "DESCRIPCION": descripcion,
                "MONTO": monto
            })
            
        return pd.DataFrame(datos_normalizados)
    except Exception as e:
        st.error(f"❌ Error procesando archivo Banplus: {e}")
        return pd.DataFrame()

# =========================================================
# 🔥 PROCESAMIENTO PRINCIPAL - CON DETECCIÓN DIRECTA PARA PROVINCIAL Y BANCAMIGA
# =========================================================

def mono_procesar_archivo(df, usar_api=False, banco=""):
    """
    Procesa el archivo y clasifica movimientos en ingresos, egresos y comisiones.
    
    Args:
        df: DataFrame con los movimientos en formato Mercantil
        usar_api: Booleano para usar API BCV
        banco: Nombre del banco (para aplicar reglas específicas)
    """
    ingresos = []
    egresos = []
    comisiones = []
    registros_procesados = set()
    
    tipos_ingresos = ["NC", "C", "CREDITO", "ABONO", "DP", "DEP"]
    tipos_egresos = ["ND", "D", "DEBITO", "DEBIT"]
    cache_tasas = {}

    for _, fila in df.iterrows():
        try:
            if len(fila) < 10:
                continue

            fecha_raw = str(fila[3]).strip()
            if fecha_raw.lower() == "nan":
                continue

            fecha_raw = fecha_raw.replace(".0", "")
            if len(fecha_raw) == 7 and fecha_raw.isdigit():
                fecha = f"0{fecha_raw[0]}/{fecha_raw[1:3]}/{fecha_raw[3:]}"
            elif len(fecha_raw) == 8 and fecha_raw.isdigit():
                fecha = f"{fecha_raw[0:2]}/{fecha_raw[2:4]}/{fecha_raw[4:]}"
            else:
                fecha = fecha_raw

            tipo = str(fila[5]).strip().upper()
            descripcion = str(fila[6]).strip()
            referencia = str(fila[4]).strip()

            monto_bs = mono_convertir_monto(fila[7])
            if monto_bs is None or monto_bs == 0:
                continue
            
            fecha_obj = pd.to_datetime(fecha, dayfirst=True, errors="coerce")
            if pd.isna(fecha_obj):
                continue
            
            fecha_key = fecha_obj.strftime("%d/%m/%Y")
            if fecha_key in cache_tasas:
                tasa = cache_tasas[fecha_key]
            else:
                tasa = obtener_tasa_por_fecha(fecha_obj, usar_api) or 1.0
                if tasa is not None:
                    cache_tasas[fecha_key] = tasa

            monto_usd = calcular_usd(monto_bs, tasa)
            if monto_usd is None:
                continue

            texto = descripcion.upper()
            if texto in ["SALDO", "DESCRIPCION", "DESCRIPCIÓN", "REFERENCIA", "MOVIMIENTO", "FECHA", "SALDO INICIAL", "SALDO FINAL"]:
                continue

            registro = {
                "FECHA": fecha,
                "REFERENCIA": referencia,
                "DESCRIPCIÓN": descripcion,
                "MONTO BS": round(abs(monto_bs), 2) if monto_bs else 0,
                "TASA BCV": round(tasa, 4),
                "MONTO USD": monto_usd,
                "STATUS": "",
                "OBSERVACIÓN": "",
                "TIPO_PAGO": "",
                "PROVEEDOR_IPAGO": "",
                "DESCRIPCION_ORIGINAL": ""
            }

            clave = (fecha, referencia, descripcion, monto_usd, tipo)
            if clave in registros_procesados:
                continue
            registros_procesados.add(clave)

            # =========================================================
            # 🔥 REGLA ESPECIAL PARA PROVINCIAL - DETECCIÓN DIRECTA
            # =========================================================
            es_comision_provincial = False
            
            if banco == "provincial":
                descripcion_upper = descripcion.upper()
                # Detectar comisiones de Provincial por "COMIS" en la descripción
                if "COMIS" in descripcion_upper:
                    es_comision_provincial = True
                    st.write(f"🔍 Comisión Provincial detectada: {descripcion} - Monto: {monto_bs}")
                
                if es_comision_provincial:
                    comisiones.append(registro)
                    continue

            # =========================================================
            # 🔥 REGLA ESPECIAL PARA BANCAMIGA - DETECCIÓN DIRECTA
            # =========================================================
            es_comision_bancamiga = False
            
            if banco == "bancamiga":
                descripcion_upper = descripcion.upper()
                # Detectar comisiones de Bancamiga por "COMISI" en la descripción
                if "COMISI" in descripcion_upper:
                    es_comision_bancamiga = True
                    st.write(f"🔍 Comisión Bancamiga detectada: {descripcion} - Monto: {monto_bs}")
                
                if es_comision_bancamiga:
                    comisiones.append(registro)
                    continue

            # =========================================================
            # 🔥 REGLA ESPECIAL PARA VENEZUELA - DETECCIÓN DIRECTA
            # =========================================================
            es_comision_bdv = False

            if banco == "venezuela":
                descripcion_upper = descripcion.upper()
                patrones_bdv = [
                    "COM PAGO OTRAS CTAS", "COMISION PAGO A PROVEEDORES", "COM PAGO OTR BCOS",
                    "COM PAGO OTRAS CTAS JUR NAT", "COM PAGO OTRAS CTAS JUR JUR", "COMISION POR TRANSFERENCIA",
                    "COMISION PAGO MOVIL", "COMISIÓN PAGO MOVIL", "COMISION X PAGO DE NOMINA",
                    "COMISION X PAGO DE NOMINAS", "ITF", "IMPUESTO A LAS TRANSACCIONES FINANCIERAS",
                    "CARGO BANCARIO", "MANTENIMIENTO DE CUENTA", "COMISION BANCARIA", "COMISIÓN BANCARIA",
                    "CARGO POR SERVICIO", "CARGO POR TRANSACCION", "COMISION PAGO MOVIL COMERCIAL",
                    "COMISION X PAGO DE NOMINAS MB", "DOMICILIACION J412438905", "DISTRIBUIDORA GLOBAL",
                    "DOMICILIACION"
                ]
                if any(p in descripcion_upper for p in patrones_bdv) or referencia.startswith(("970", "972", "067")) or (tipo == "ND" and "COM" in descripcion_upper):
                    es_comision_bdv = True

                if es_comision_bdv:
                    comisiones.append(registro)
                    continue

            # =========================================================
            # 🔥 REGLA ESPECIAL PARA MERCANTIL - TODAS LAS COMISIONES
            # =========================================================
            es_comision_mercantil = False
            
            if banco == "mercantil":
                descripcion_upper = descripcion.upper()
                
                # 🔥 Lista completa de patrones de comisiones de Mercantil
                patrones_comision_mercantil = [
                    "OP.CRED.DIRT. CLTE-CLTE",
                    "OP CRED DIRT CLTE CLTE",
                    "COMISION PAGO MOVIL COMERCIAL",
                    "COMISION POR TRANSFERENCIA DE FONDOS",
                    "COMISION X PAGO DE NOMINAS",
                    "COMISION PAGO MOVIL COMERCIAL INTERBANCARIO",
                    "COMISION X PAGO DE NOMINAS MB",
                    "ITF",
                    "IMPUESTO A LAS TRANSACCIONES FINANCIERAS",
                    "CARGO BANCARIO",
                    "MANTENIMIENTO DE CUENTA",
                    "COMISION POR TRANSFERENCIA",
                    "EMISION EDO",
                    "RETENCION DE IMPUESTO"
                ]
                
                for patron in patrones_comision_mercantil:
                    if patron in descripcion_upper:
                        es_comision_mercantil = True
                        break
                
                # Si es comisión de Mercantil, la clasificamos como tal
                if es_comision_mercantil:
                    comisiones.append(registro)
                    continue

            # =========================================================
            # 🔥 REGLA ESPECIAL PARA TESORO - DETECCIÓN DIRECTA
            # =========================================================
            es_comision_tesoro = False
            
            if banco == "tesoro":
                descripcion_upper = descripcion.upper()
                patrones_tesoro = [
                    "BELOW MINIMUM BALANCE CHARGES",
                    "STAMENT SERVICE",
                    "STATEMENT SERVICE",
                    "COMIS",
                    "COMISION",
                    "CARGO BANCARIO",
                    "CARGO POR SERVICIO"
                ]
                for patron in patrones_tesoro:
                    if patron in descripcion_upper:
                        es_comision_tesoro = True
                        break
                
                if es_comision_tesoro:
                    comisiones.append(registro)
                    continue

            # 🔥 PASAR EL PROVEEDOR A LA FUNCIÓN es_comision
            proveedor = fila.get("Proveedor") if isinstance(fila, dict) else None
            if mono_es_comision(descripcion, proveedor):
                comisiones.append(registro)
            elif tipo in tipos_ingresos:
                ingresos.append(registro)
            elif tipo in tipos_egresos:
                egresos.append(registro)

        except Exception as e:
            continue

    return ingresos, egresos, comisiones

# =========================================================

# =========================================================
# HEADER
# =========================================================

col_l, col_c, col_r = st.columns([2, 1, 2])

with col_c:
    try:
        st.image("LOGO.jpeg", width=145)
    except Exception:
        st.image(
            "https://raw.githubusercontent.com/pelobravo/clasificador-excel/main/LOGO.jpeg",
            width=145
        )

st.markdown("<h1 style='text-align: center; color: #1e3a5f; margin-top: 15px; margin-bottom: 5px;'>Clasificador Bancario</h1>", unsafe_allow_html=True)
st.markdown("<h3 style='text-align: center; color: #00a8cc; margin-top: 0px; margin-bottom: 20px; font-weight: 600;'>Grupo Bodeguita Oriente</h3>", unsafe_allow_html=True)
st.markdown("---")

# Inicializar estado de navegación
if "seccion_activa" not in st.session_state:
    st.session_state.seccion_activa = "consolidado"

# Barra de Navegación Superior
col_nav1, col_nav2 = st.columns(2)
with col_nav1:
    if st.button("📊 SALDOS BANCARIOS MULTIBANCO", use_container_width=True, type="primary" if st.session_state.seccion_activa == "consolidado" else "secondary", key="nav_btn_consolidado"):
        st.session_state.seccion_activa = "consolidado"
        st.rerun()
with col_nav2:
    if st.button("🔍 CRUCE DE INFORMACIÓN (MONOBANCO)", use_container_width=True, type="primary" if st.session_state.seccion_activa == "cruce" else "secondary", key="nav_btn_cruce"):
        st.session_state.seccion_activa = "cruce"
        st.rerun()
st.markdown("---")

# =========================================================
# SIDEBAR DINÁMICO Y DECLARACIÓN DE VARIABLES
# =========================================================

# Declaración de variables para evitar NameError en cualquiera de los dos flujos
archivo_ipago = None
archivo_banesco = None
archivo_bnc = None
archivo_mercantil = None
archivo_venezuela = None
archivo_provincial = None
archivo_bancamiga = None
saldo_manual_tesoro = 0.0

archivo = None

with st.sidebar:
    st.markdown(
        """
        <div style="display: flex; justify-content: center; align-items: center; background-color: #f1f3f5; padding: 12px; border-radius: 10px; border: 1px solid #e9ecef; margin-bottom: 5px;">
            <img src="https://raw.githubusercontent.com/pelobravo/clasificador-excel/main/LOGO.jpeg" style="width: 100px; border-radius: 8px;">
        </div>
        """,
        unsafe_allow_html=True
    )
    st.markdown("---")

    if st.session_state.get("seccion_activa", "consolidado") == "consolidado":
        st.markdown("### 📂 Cargar Archivos (Consolidado)")

        with st.expander("📊 iPago (Archivo Maestro)", expanded=True):
            archivo_ipago = st.file_uploader(
                "Cargar archivo iPago",
                type=["xlsx", "xls", "xlsm"],
                key="uploader_ipago"
            )

        st.markdown("#### 🏦 Bancos")

        with st.expander("🏦 Banesco", expanded=False):
            archivo_banesco = st.file_uploader(
                "Cargar Banesco",
                type=["xlsx", "xls", "xlsm"],
                accept_multiple_files=True,
                key="uploader_banesco"
            )

        with st.expander("🏦 BNC", expanded=False):
            archivo_bnc = st.file_uploader(
                "Cargar BNC",
                type=["xlsx", "xls", "xlsm"],
                accept_multiple_files=True,
                key="uploader_bnc"
            )

        with st.expander("🏦 Mercantil", expanded=False):
            archivo_mercantil = st.file_uploader(
                "Cargar Mercantil",
                type=["xlsx", "xls", "xlsm"],
                accept_multiple_files=True,
                key="uploader_mercantil"
            )

        with st.expander("🏦 Banco de Venezuela (BDV)", expanded=False):
            archivo_venezuela = st.file_uploader(
                "Cargar BDV",
                type=["xlsx", "xls", "xlsm"],
                accept_multiple_files=True,
                key="uploader_venezuela"
            )

        with st.expander("🏦 Provincial", expanded=False):
            archivo_provincial = st.file_uploader(
                "Cargar Provincial",
                type=["xlsx", "xls", "xlsm"],
                accept_multiple_files=True,
                key="uploader_provincial"
            )

        with st.expander("🏦 Bancamiga", expanded=False):
            archivo_bancamiga = st.file_uploader(
                "Cargar Bancamiga",
                type=["xlsx", "xls", "xlsm"],
                accept_multiple_files=True,
                key="uploader_bancamiga"
            )

        with st.expander("🏦 BanPlus", expanded=False):
            archivo_banplus = st.file_uploader(
                "Cargar BanPlus",
                type=["xlsx", "xls", "xlsm"],
                accept_multiple_files=True,
                key="uploader_banplus"
            )

        with st.expander("🏦 Banco del Tesoro", expanded=False):
            saldo_manual_tesoro = st.number_input(
                "Saldo manual Banco del Tesoro (VES)",
                min_value=0.0,
                value=0.0,
                step=100.0,
                key="saldo_manual_tesoro"
            )

        with st.expander("💵 Banco Efectivo (Manual)", expanded=False):
            saldo_manual_efectivo = st.number_input(
                "Saldo manual Banco Efectivo (USD)",
                min_value=0.0,
                value=0.0,
                step=100.0,
                key="saldo_manual_efectivo"
            )

        with st.expander("🪙 Banco Binance (Manual)", expanded=False):
            saldo_manual_binance = st.number_input(
                "Saldo manual Banco Binance (USD)",
                min_value=0.0,
                value=0.0,
                step=100.0,
                key="saldo_manual_binance"
            )
    else:
        st.markdown("### 📂 Cargar Archivo Único (Monobanco)")

        archivo = st.file_uploader(
            "📂 Cargar archivo Excel",
            type=["xlsx", "xls", "xlsm"],
            key="uploader_monobanco"
        )

        archivo_ipago = st.file_uploader(
            "📂 Cargar archivo iPago",
            type=["xlsx", "xls", "xlsm"],
            key="uploader_ipago_mono"
        )

    st.markdown("---")

    fecha_inicio = st.date_input(
        "📅 Fecha Inicio",
        value=date.today().replace(day=1)
    )

    fecha_fin = st.date_input(
        "📅 Fecha Fin",
        value=date.today()
    )

    st.markdown("---")

    usar_api = st.checkbox(
        "🌐 Usar API BCV automática (experimental)",
        value=False
    )

    st.markdown("---")

    procesar = st.button(
        "🚀 Procesar",
        type="primary",
        use_container_width=True
    )

# =========================================================
# MAIN BODY CONDITIONAL ROUTING
# =========================================================

if st.session_state.seccion_activa == "consolidado":
    # ---------------------------------------------------------
    # FLUX 1: CIERRE CONSOLIDADO MULTIBANCO
    # ---------------------------------------------------------
    if not archivo_banesco: st.session_state.saldo_banesco = 0.0
    if not archivo_bnc: st.session_state.saldo_bnc = 0.0
    if not archivo_mercantil: st.session_state.saldo_mercantil = 0.0
    if not archivo_venezuela: st.session_state.saldo_venezuela = 0.0
    if not archivo_provincial: st.session_state.saldo_provincial = 0.0
    if not archivo_bancamiga: st.session_state.saldo_bancamiga = 0.0
    if not archivo_banplus: st.session_state.saldo_banplus = 0.0
    st.session_state.saldo_tesoro = st.session_state.get("saldo_manual_tesoro", 0.0)
    st.session_state.saldo_efectivo = st.session_state.get("saldo_manual_efectivo", 0.0)
    st.session_state.saldo_binance = st.session_state.get("saldo_manual_binance", 0.0)

    # Selector de Moneda para los KPIs
    col_mon1, col_mon2 = st.columns([1.5, 3.5])
    with col_mon1:
        st.radio(
            "💱 Mostrar KPIs en:",
            options=["Dólares ($)", "Bolívares (Bs.)"],
            horizontal=True,
            key="selector_moneda_kpis"
        )
    
    # Determinar moneda seleccionada
    moneda_kpi = "USD" if st.session_state.get("selector_moneda_kpis", "Dólares ($)") == "Dólares ($)" else "VES"

    # Renderizado de KPIs
    tasa_dia = obtener_tasa_bcv(fecha_fin, usar_api)
    
    # Efectivo y Binance se ingresan en USD, los convertimos a VES usando la tasa del día
    st.session_state.saldo_efectivo = st.session_state.get("saldo_manual_efectivo", 0.0) * tasa_dia
    st.session_state.saldo_binance = st.session_state.get("saldo_manual_binance", 0.0) * tasa_dia

    total_ves = (
        st.session_state.saldo_banesco + st.session_state.saldo_bnc + 
        st.session_state.saldo_mercantil + st.session_state.saldo_venezuela + 
        st.session_state.saldo_provincial + st.session_state.saldo_bancamiga + 
        st.session_state.saldo_banplus + 
        st.session_state.saldo_tesoro +
        st.session_state.saldo_efectivo +
        st.session_state.saldo_binance
    )
    total_usd = total_ves / tasa_dia if tasa_dia > 0 else 0.0
    
    total_ingresos_ves = st.session_state.get("total_ingresos_consolidado", 0.0)
    total_ingresos_usd = total_ingresos_ves / tasa_dia if tasa_dia > 0 else 0.0
    
    total_egresos_ves = st.session_state.get("total_egresos_ipago_ves", 0.0)
    total_egresos_usd = total_egresos_ves / tasa_dia if tasa_dia > 0 else 0.0

    bancos_con_saldo = []
    if st.session_state.saldo_banesco > 0: bancos_con_saldo.append(f"Banesco: Bs. {formato_venezolano(st.session_state.saldo_banesco)}")
    if st.session_state.saldo_bnc > 0: bancos_con_saldo.append(f"BNC: Bs. {formato_venezolano(st.session_state.saldo_bnc)}")
    if st.session_state.saldo_mercantil > 0: bancos_con_saldo.append(f"Mercantil: Bs. {formato_venezolano(st.session_state.saldo_mercantil)}")
    if st.session_state.saldo_venezuela > 0: bancos_con_saldo.append(f"BDV: Bs. {formato_venezolano(st.session_state.saldo_venezuela)}")
    if st.session_state.saldo_provincial > 0: bancos_con_saldo.append(f"Provincial: Bs. {formato_venezolano(st.session_state.saldo_provincial)}")
    if st.session_state.saldo_bancamiga > 0: bancos_con_saldo.append(f"Bancamiga: Bs. {formato_venezolano(st.session_state.saldo_bancamiga)}")
    if st.session_state.saldo_banplus > 0: bancos_con_saldo.append(f"BanPlus: Bs. {formato_venezolano(st.session_state.saldo_banplus)}")
    if st.session_state.saldo_tesoro > 0: bancos_con_saldo.append(f"Tesoro: Bs. {formato_venezolano(st.session_state.saldo_tesoro)}")
    saldo_ef_usd = st.session_state.get("saldo_manual_efectivo", 0.0)
    if st.session_state.saldo_efectivo > 0: bancos_con_saldo.append(f"Efectivo: Bs. {formato_venezolano(st.session_state.saldo_efectivo)} (${saldo_ef_usd:,.2f})")
    saldo_bin_usd = st.session_state.get("saldo_manual_binance", 0.0)
    if st.session_state.saldo_binance > 0: bancos_con_saldo.append(f"Binance: Bs. {formato_venezolano(st.session_state.saldo_binance)} (${saldo_bin_usd:,.2f})")

    kpi_subtitle_text = " | ".join(bancos_con_saldo) if bancos_con_saldo else "Sin saldos cargados"

    if moneda_kpi == "USD":
        val_saldos = f"${total_usd:,.2f}"
        sub_saldos = f"Bs. {formato_venezolano(total_ves)}"
        
        val_ingresos = f"${total_ingresos_usd:,.2f}"
        sub_ingresos = f"Bs. {formato_venezolano(total_ingresos_ves)}"
        
        val_egresos = f"${total_egresos_usd:,.2f}"
        sub_egresos = f"Bs. {formato_venezolano(total_egresos_ves)}"
        
        title_saldos = "Total Saldos Bancos (USD)"
        title_ingresos = "Total Ingresos Bancos (USD)"
        title_egresos = "Total Egresos iPago (USD)"
        
        title_extra = "Total Equivalente (VES)"
        val_extra = f"Bs. {formato_venezolano(total_ves)}"
        sub_extra = "Saldos convertidos a tasa oficial"
    else:
        val_saldos = f"Bs. {formato_venezolano(total_ves)}"
        sub_saldos = f"${total_usd:,.2f} USD"
        
        val_ingresos = f"Bs. {formato_venezolano(total_ingresos_ves)}"
        sub_ingresos = f"${total_ingresos_usd:,.2f} USD"
        
        val_egresos = f"Bs. {formato_venezolano(total_egresos_ves)}"
        sub_egresos = f"${total_egresos_usd:,.2f} USD"
        
        title_saldos = "Total Saldos Bancos (VES)"
        title_ingresos = "Total Ingresos Bancos (VES)"
        title_egresos = "Total Egresos iPago (VES)"
        
        title_extra = "Total Equivalente (USD)"
        val_extra = f"${total_usd:,.2f}"
        sub_extra = "Convertido al tipo de cambio oficial"

    st.markdown(f"""
    <div class="kpi-container">
        <div class="kpi-card">
            <div class="kpi-title">{title_saldos}</div>
            <div class="kpi-value">{val_saldos}</div>
            <div class="kpi-subtitle">{kpi_subtitle_text if moneda_kpi == 'VES' else sub_saldos}</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-title">{title_ingresos}</div>
            <div class="kpi-value">{val_ingresos}</div>
            <div class="kpi-subtitle">{sub_ingresos}</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-title">{title_egresos}</div>
            <div class="kpi-value">{val_egresos}</div>
            <div class="kpi-subtitle">{sub_egresos}</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-title">Tasa Oficial BCV del Día</div>
            <div class="kpi-value">{tasa_dia:.4f} VES/USD</div>
            <div class="kpi-subtitle">Tasa del Banco Central de Venezuela</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-title">{title_extra}</div>
            <div class="kpi-value">{val_extra}</div>
            <div class="kpi-subtitle">{sub_extra}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # =========================================================
    # LEER ARCHIVOS CARGADOS
    # =========================================================
    df_ipago = None
    if archivo_ipago:
        try:
            df_ipago = pd.read_excel(archivo_ipago, engine="openpyxl")
            df_ipago.columns = [str(c).strip() for c in df_ipago.columns]
            st.success(f"✅ Archivo iPago cargado: {len(df_ipago)} registros")
        except Exception as e:
            st.error(f"❌ Error leyendo archivo iPago: {e}")

    list_df_convertidos = []
    bancos_procesados = []
    saldos_detalle_excel = []

    # 1. Banesco
    if archivo_banesco:
        st.session_state.saldo_banesco = 0.0
        for idx, arch in enumerate(archivo_banesco, 1):
            try:
                nombre = arch.name.lower()
                if nombre.endswith(".xlsx") or nombre.endswith(".xlsm"):
                    df_raw = pd.read_excel(arch, engine="openpyxl", header=None)
                else:
                    df_raw = pd.read_html(arch)[0]
            
                saldo_arch = obtener_saldo_banco(df_raw, "banesco")
                st.session_state.saldo_banesco += saldo_arch
            
                nombre_banco = f"Banesco - Cuenta {idx}" if len(archivo_banesco) > 1 else "Banesco"
                saldos_detalle_excel.append((nombre_banco, saldo_arch))
            
                df_normalizado = procesar_banesco(df_raw)
                df_convertido = convertir_a_formato_mercantil(df_normalizado, "banesco")
                if not df_convertido.empty:
                    list_df_convertidos.append(df_convertido)
                    if "Banesco" not in bancos_procesados:
                        bancos_procesados.append("Banesco")
            except Exception as e:
                st.error(f"❌ Error leyendo Banesco ({arch.name}): {e}")
    else:
        saldos_detalle_excel.append(("Banesco", 0.0))

    # 2. BNC
    if archivo_bnc:
        st.session_state.saldo_bnc = 0.0
        for idx, arch in enumerate(archivo_bnc, 1):
            try:
                df_raw = leer_excel_con_encabezados(arch)
                encabezado = None
                for i in range(min(30, len(df_raw))):
                    fila = df_raw.iloc[i].fillna("").astype(str)
                    texto = " ".join(fila.tolist()).lower()
                    if "fecha" in texto and ("descripcion" in texto or "descripción" in texto):
                        encabezado = i
                        break
            
                saldo_arch = obtener_saldo_banco(df_raw, "bnc", encabezado)
                st.session_state.saldo_bnc += saldo_arch
            
                nombre_banco = f"BNC - Cuenta {idx}" if len(archivo_bnc) > 1 else "BNC"
                saldos_detalle_excel.append((nombre_banco, saldo_arch))
            
                df_normalizado = procesar_bnc(df_raw)
                df_convertido = convertir_a_formato_mercantil(df_normalizado, "bnc")
                if not df_convertido.empty:
                    list_df_convertidos.append(df_convertido)
                    if "BNC" not in bancos_procesados:
                        bancos_procesados.append("BNC")
            except Exception as e:
                st.error(f"❌ Error leyendo BNC ({arch.name}): {e}")
    else:
        saldos_detalle_excel.append(("BNC", 0.0))

    # 3. Mercantil
    if archivo_mercantil:
        st.session_state.saldo_mercantil = 0.0
        for idx, arch in enumerate(archivo_mercantil, 1):
            try:
                df_raw = leer_excel_sin_encabezados(arch)
                df_raw = preparar_df_con_encabezado_dinamico(df_raw)
                saldo_arch = obtener_saldo_banco(df_raw, "mercantil")
                st.session_state.saldo_mercantil += saldo_arch
            
                nombre_banco = f"Mercantil - Cuenta {idx}" if len(archivo_mercantil) > 1 else "Mercantil"
                saldos_detalle_excel.append((nombre_banco, saldo_arch))
            
                df_convertido = convertir_a_formato_mercantil(df_raw, "mercantil")
                if not df_convertido.empty:
                    list_df_convertidos.append(df_convertido)
                    if "Mercantil" not in bancos_procesados:
                        bancos_procesados.append("Mercantil")
            except Exception as e:
                st.error(f"❌ Error leyendo Mercantil ({arch.name}): {e}")
    else:
        saldos_detalle_excel.append(("Mercantil", 0.0))

    # 4. BDV
    if archivo_venezuela:
        st.session_state.saldo_venezuela = 0.0
        for idx, arch in enumerate(archivo_venezuela, 1):
            try:
                df_raw = leer_excel_sin_encabezados(arch)
                saldo_arch = obtener_saldo_banco(df_raw, "venezuela")
                st.session_state.saldo_venezuela += saldo_arch
            
                nombre_banco = f"Banco de Venezuela (BDV) - Cuenta {idx}" if len(archivo_venezuela) > 1 else "Banco de Venezuela (BDV)"
                saldos_detalle_excel.append((nombre_banco, saldo_arch))
            
                df_normalizado = procesar_venezuela_simple(df_raw)
                df_convertido = convertir_venezuela_a_formato_mercantil(df_normalizado)
                if not df_convertido.empty:
                    list_df_convertidos.append(df_convertido)
                    if "Venezuela" not in bancos_procesados:
                        bancos_procesados.append("Venezuela")
            except Exception as e:
                st.error(f"❌ Error leyendo BDV ({arch.name}): {e}")
    else:
        saldos_detalle_excel.append(("Banco de Venezuela (BDV)", 0.0))

    # 5. Provincial
    if archivo_provincial:
        st.session_state.saldo_provincial = 0.0
        for idx, arch in enumerate(archivo_provincial, 1):
            try:
                df_raw = leer_excel_sin_encabezados(arch)
                saldo_arch = obtener_saldo_banco(df_raw, "provincial")
                st.session_state.saldo_provincial += saldo_arch
            
                nombre_banco = f"Provincial - Cuenta {idx}" if len(archivo_provincial) > 1 else "Provincial"
                saldos_detalle_excel.append((nombre_banco, saldo_arch))
            
                df_normalizado = procesar_provincial(df_raw)
                df_convertido = convertir_a_formato_mercantil(df_normalizado, "provincial")
                if not df_convertido.empty:
                    list_df_convertidos.append(df_convertido)
                    if "Provincial" not in bancos_procesados:
                        bancos_procesados.append("Provincial")
            except Exception as e:
                st.error(f"❌ Error leyendo Provincial ({arch.name}): {e}")
    else:
        saldos_detalle_excel.append(("Provincial", 0.0))

    # 6. Bancamiga
    if archivo_bancamiga:
        st.session_state.saldo_bancamiga = 0.0
        for idx, arch in enumerate(archivo_bancamiga, 1):
            try:
                nombre = arch.name.lower()
                if nombre.endswith(".xlsx") or nombre.endswith(".xlsm"):
                    df_raw = pd.read_excel(arch, engine="openpyxl", header=None)
                else:
                    try:
                        df_raw = pd.read_excel(arch, header=None)
                    except Exception:
                        arch.seek(0)
                        df_raw = pd.read_html(arch, decimal=',', thousands='.')[0]
                
                if isinstance(df_raw.columns, pd.MultiIndex):
                    df_raw.columns = df_raw.columns.get_level_values(-1)
            
                saldo_arch = obtener_saldo_banco(df_raw, "bancamiga")
                st.session_state.saldo_bancamiga += saldo_arch
            
                nombre_banco = f"Bancamiga - Cuenta {idx}" if len(archivo_bancamiga) > 1 else "Bancamiga"
                saldos_detalle_excel.append((nombre_banco, saldo_arch))
            
                df_normalizado = procesar_bancamiga(df_raw)
                df_convertido = convertir_a_formato_mercantil(df_normalizado, "bancamiga")
                if not df_convertido.empty:
                    list_df_convertidos.append(df_convertido)
                    if "Bancamiga" not in bancos_procesados:
                        bancos_procesados.append("Bancamiga")
            except Exception as e:
                st.error(f"❌ Error leyendo Bancamiga ({arch.name}): {e}")
    else:
        saldos_detalle_excel.append(("Bancamiga", 0.0))

    # 6.5. BanPlus
    if archivo_banplus:
        st.session_state.saldo_banplus = 0.0
        for idx, arch in enumerate(archivo_banplus, 1):
            try:
                nombre = arch.name.lower()
                if nombre.endswith(".xlsx") or nombre.endswith(".xlsm"):
                    df_raw = pd.read_excel(arch, engine="openpyxl", header=None)
                else:
                    try:
                        df_raw = pd.read_excel(arch, header=None)
                    except Exception:
                        arch.seek(0)
                        df_raw = pd.read_html(arch)[0]
            
                saldo_arch = obtener_saldo_banco(df_raw, "banplus")
                st.session_state.saldo_banplus += saldo_arch
            
                nombre_banco = f"BanPlus - Cuenta {idx}" if len(archivo_banplus) > 1 else "BanPlus"
                saldos_detalle_excel.append((nombre_banco, saldo_arch))
            
                df_normalizado = procesar_banplus(df_raw)
                df_convertido = convertir_a_formato_mercantil(df_normalizado, "banplus")
                if not df_convertido.empty:
                    list_df_convertidos.append(df_convertido)
                    if "BanPlus" not in bancos_procesados:
                        bancos_procesados.append("BanPlus")
            except Exception as e:
                st.error(f"❌ Error leyendo BanPlus ({arch.name}): {e}")
    else:
        saldos_detalle_excel.append(("BanPlus", 0.0))

    # 7. Tesoro (Manual)
    saldos_detalle_excel.append(("Banco del Tesoro", st.session_state.saldo_tesoro))

    # 7.5. Efectivo (Manual)
    saldos_detalle_excel.append(("Banco Efectivo", st.session_state.saldo_efectivo))

    # 7.6. Binance (Manual)
    saldos_detalle_excel.append(("Banco Binance", st.session_state.saldo_binance))

    st.session_state.saldos_detalle_excel = saldos_detalle_excel

    # Recalcular el total consolidado con los datos extraídos
    st.session_state.saldo_efectivo = st.session_state.get("saldo_manual_efectivo", 0.0) * tasa_dia
    st.session_state.saldo_binance = st.session_state.get("saldo_manual_binance", 0.0) * tasa_dia

    total_ves = (
        st.session_state.saldo_banesco + st.session_state.saldo_bnc + 
        st.session_state.saldo_mercantil + st.session_state.saldo_venezuela + 
        st.session_state.saldo_provincial + st.session_state.saldo_bancamiga + 
        st.session_state.saldo_banplus + 
        st.session_state.saldo_tesoro +
        st.session_state.saldo_efectivo +
        st.session_state.saldo_binance
    )
    total_usd = total_ves / tasa_dia if tasa_dia > 0 else 0.0

    if list_df_convertidos:
        df_original = pd.concat(list_df_convertidos, ignore_index=True)
    
        # Filtrar por fechas
        try:
            def parsear_fechas_consolidado(columna_fechas):
                fechas = []
                for val in columna_fechas:
                    val_str = str(val).strip().replace(".0", "")
                    if not val_str or val_str == "nan":
                        fechas.append(pd.NaT)
                        continue
                
                    # Si es numérico de 8 dígitos (formato Mercantil ddmmyyyy)
                    if len(val_str) == 8 and val_str.isdigit():
                        dt = pd.to_datetime(val_str, format="%d%m%Y", errors="coerce")
                        if pd.notna(dt):
                            fechas.append(dt)
                            continue
                
                    # Parseo flexible general
                    parsed = False
                    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"]:
                        dt = pd.to_datetime(val_str, format=fmt, errors="coerce")
                        if pd.notna(dt):
                            fechas.append(dt)
                            parsed = True
                            break
                    if not parsed:
                        dt = pd.to_datetime(val_str, errors="coerce", dayfirst=True)
                        fechas.append(dt)
                return pd.Series(fechas)
            
            fechas_convertidas = parsear_fechas_consolidado(df_original.iloc[:, 3])
            fecha_inicio_dt = pd.to_datetime(fecha_inicio)
            fecha_fin_dt = pd.to_datetime(fecha_fin)
            
            # Ajustar automáticamente el rango si las fechas consolidadas están fuera del rango seleccionado
            if not fechas_convertidas.empty:
                min_file_date = fechas_convertidas.min()
                max_file_date = fechas_convertidas.max()
                if pd.notna(min_file_date) and pd.notna(max_file_date):
                    if fecha_inicio_dt > min_file_date or fecha_fin_dt < max_file_date:
                        fecha_inicio_dt = min_file_date
                        fecha_fin_dt = max_file_date
                        st.info(f"💡 **Rango de fechas ajustado automáticamente** al contenido de los archivos: {min_file_date.strftime('%d/%m/%Y')} al {max_file_date.strftime('%d/%m/%Y')}")
            
            df_original = df_original[(fechas_convertidas >= fecha_inicio_dt) & (fechas_convertidas <= fecha_fin_dt)]
            
            # Calcular la suma de ingresos de los archivos consolidados
            total_ingresos_ves_calc = 0.0
            if not df_original.empty:
                tipos_ingresos = ["NC", "C", "CREDITO", "ABONO", "DP", "DEP"]
                tipo_col = df_original.iloc[:, 5].astype(str).str.strip().str.upper()
                ingresos_filas = df_original[tipo_col.isin(tipos_ingresos)]
                
                for val in ingresos_filas.iloc[:, 7]:
                    try:
                        if isinstance(val, (int, float)):
                            total_ingresos_ves_calc += abs(float(val))
                        else:
                            val_str = str(val).strip()
                            if "," in val_str and "." in val_str:
                                if val_str.find(".") < val_str.find(","):
                                    val_str = val_str.replace(".", "").replace(",", ".")
                                else:
                                    val_str = val_str.replace(",", "")
                            elif "," in val_str:
                                val_str = val_str.replace(",", ".")
                            total_ingresos_ves_calc += abs(float(val_str))
                    except:
                        pass
            st.session_state.total_ingresos_consolidado = total_ingresos_ves_calc
            
            # Calcular la suma de egresos del archivo iPago
            total_egresos_ipago_ves_calc = 0.0
            if df_ipago is not None and not df_ipago.empty:
                try:
                    fechas_ipago = pd.to_datetime(df_ipago['Fecha Pago'], errors='coerce')
                    df_ipago_filtrado = df_ipago[(fechas_ipago >= fecha_inicio_dt) & (fechas_ipago <= fecha_fin_dt)]
                    # Sumar la columna Monto convirtiendo a float por seguridad
                    monto_sum = 0.0
                    for val in df_ipago_filtrado['Monto']:
                        try:
                            if isinstance(val, (int, float)):
                                monto_sum += abs(float(val))
                            else:
                                val_str = str(val).strip()
                                if "," in val_str and "." in val_str:
                                    if val_str.find(".") < val_str.find(","):
                                        val_str = val_str.replace(".", "").replace(",", ".")
                                    else:
                                        val_str = val_str.replace(",", "")
                                elif "," in val_str:
                                    val_str = val_str.replace(",", ".")
                                monto_sum += abs(float(val_str))
                        except:
                            pass
                    total_egresos_ipago_ves_calc = monto_sum
                except Exception as e:
                    st.warning(f"⚠️ Error al calcular egresos de iPago: {e}")
            st.session_state.total_egresos_ipago_ves = total_egresos_ipago_ves_calc
            
            st.success(f"📅 Movimientos consolidados de {', '.join(bancos_procesados)} filtrados del {fecha_inicio_dt.strftime('%d/%m/%Y')} al {fecha_fin_dt.strftime('%d/%m/%Y')} ({len(df_original)} registros)")
        except Exception as e:
            st.warning(f"⚠️ Error filtrando fechas consolidadas: {e}")

        if df_original.empty:
            st.warning("⚠️ No se encontraron movimientos en el rango de fechas.")
        else:
            with st.expander("👁️ Vista previa de movimientos consolidados (Formato Unificado)"):
                st.dataframe(df_original.head(20), use_container_width=True)

            if procesar:
                with st.spinner("🚀 Conciliando y cruzando transacciones con iPago..."):
                    ingresos, egresos, comisiones = procesar_archivo(df_original, usar_api, banco="multibanco")
                    df_ingresos = pd.DataFrame(ingresos)
                    df_egresos = pd.DataFrame(egresos)
                    df_comisiones = pd.DataFrame(comisiones)

                    # Cruce con iPago
                    if df_ipago is not None and not df_egresos.empty:
                        df_egresos = enriquecer_egresos_con_ipago(df_egresos, df_ipago)
                        if "ES_COMISION" in df_egresos.columns:
                            mascara = df_egresos["ES_COMISION"] == True
                            if mascara.any():
                                df_comisiones_extra = df_egresos[mascara].copy().drop(columns=["ES_COMISION", "REFERENCIA_IPAGO"], errors="ignore")
                                df_comisiones = pd.concat([df_comisiones, df_comisiones_extra], ignore_index=True) if not df_comisiones.empty else df_comisiones_extra
                                df_egresos = df_egresos[~mascara].copy()
                                st.success(f"💳 Se identificaron {len(df_comisiones_extra)} comisiones adicionales vía iPago.")
                        df_egresos = df_egresos.drop(columns=["ES_COMISION", "REFERENCIA_IPAGO"], errors="ignore")
                        st.success(f"🎯 Cruce completado. Egresos conciliados con iPago: {len(df_egresos)} registros.")

                    for df_t in [df_ingresos, df_egresos, df_comisiones]:
                        if not df_t.empty:
                            for col in ["STATUS", "OBSERVACIÓN", "TIPO_PAGO", "PROVEEDOR_IPAGO", "DESCRIPCION_ORIGINAL"]:
                                if col not in df_t.columns: df_t[col] = ""

                    total_ingresos = df_ingresos["MONTO USD"].sum() if not df_ingresos.empty else 0
                    total_egresos = df_egresos["MONTO USD"].sum() if not df_egresos.empty else 0
                    total_comisiones = df_comisiones["MONTO USD"].sum() if not df_comisiones.empty else 0
                    neto_procesado = total_ingresos - total_egresos - total_comisiones

                    # Mostrar métricas
                    col1_m, col2_m, col3_m, col4_m = st.columns(4)
                    with col1_m: st.metric("💰 TOTAL INGRESOS PROCESADOS", len(df_ingresos), f"${total_ingresos:,.2f}")
                    with col2_m: st.metric("💸 TOTAL EGRESOS PROCESADOS", len(df_egresos), f"${total_egresos:,.2f}")
                    with col3_m: st.metric("💳 TOTAL COMISIONES PROCESADAS", len(df_comisiones), f"${total_comisiones:,.2f}")
                    with col4_m: st.metric("⚖️ NETO PROCESADO (ING - EGR - COM)", "", f"${neto_procesado:,.2f}")

                    st.subheader("📊 Detalle de Transacciones Conciliadas")
                    tab1, tab2, tab3 = st.tabs(["📈 INGRESOS", "📉 EGRESOS", "💳 COMISIONES"])
                    with tab1: st.dataframe(df_ingresos, use_container_width=True)
                    with tab2: st.dataframe(df_egresos, use_container_width=True)
                    with tab3: st.dataframe(df_comisiones, use_container_width=True)

                    # =========================================================
                    # MOTOR DE REPORTES EXCEL OPENPYXL COMPLETO
                    # =========================================================
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        workbook = writer.book
                    
                        # -----------------------------------------------------
                        # PESTAÑA: RESUMEN DE SALDOS CONSOLIDADO
                        # -----------------------------------------------------
                        hoja_resumen = workbook.create_sheet(title="RESUMEN", index=0)
                    
                        rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        azul_oscuro = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                        verde_claro = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                        gris_claro = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        amarillo = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                        blanco_bold = Font(color="FFFFFF", bold=True, size=11)
                        negro_bold = Font(color="000000", bold=True, size=11)
                        borde_fino = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
                        alineacion_centro = Alignment(horizontal="center", vertical="center")
                        centro = alineacion_centro
                        alineacion_derecha = Alignment(horizontal="right", vertical="center")
                        alineacion_izquierda = Alignment(horizontal="left", vertical="center")

                        # Cabecera Resumen (A la izquierda)
                        hoja_resumen["B2"] = "GRUPO BODEGUITA ORIENTE"
                        hoja_resumen["B2"].font = Font(bold=True, size=14, color="1E3A5F")
                        hoja_resumen["B2"].alignment = alineacion_izquierda

                        hoja_resumen["B3"] = "CONCILIACIÓN BANCARIA - RESUMEN DE SALDOS"
                        hoja_resumen["B3"].font = Font(bold=True, size=11, color="555555")
                        hoja_resumen["B3"].alignment = alineacion_izquierda

                        # Datos de la Empresa / Reporte (A la derecha superior)
                        hoja_resumen["E2"] = "Fecha de Cierre:"
                        hoja_resumen["E2"].font = Font(bold=True, size=11, color="1E3A5F")
                        hoja_resumen["E2"].alignment = alineacion_derecha
                    
                        hoja_resumen["F2"] = date.today().strftime("%d/%m/%Y")
                        hoja_resumen["F2"].font = Font(bold=False, size=11)
                        hoja_resumen["F2"].alignment = alineacion_izquierda

                        hoja_resumen["E3"] = "Tasa BCV del Día:"
                        hoja_resumen["E3"].font = Font(bold=True, size=11, color="1E3A5F")
                        hoja_resumen["E3"].alignment = alineacion_derecha
                    
                        hoja_resumen["F3"] = tasa_dia
                        hoja_resumen["F3"].font = Font(bold=False, size=11)
                        hoja_resumen["F3"].number_format = '#,##0.0000'
                        hoja_resumen["F3"].alignment = alineacion_izquierda

                        # Cabeceras tabla (3 columnas: BANCOS, TOTAL (VES), CONVERSIÓN (USD))
                        headers_r = ["BANCOS", "TOTAL (VES)", "CONVERSIÓN (USD)"]
                        for col_num, header in enumerate(headers_r, 2):
                            cell = hoja_resumen.cell(row=8, column=col_num)
                            cell.value = header
                            cell.fill = azul_oscuro
                            cell.font = blanco_bold
                            cell.alignment = alineacion_centro
                            cell.border = borde_fino

                        # Datos
                        bancos_data = st.session_state.get("saldos_detalle_excel", [
                            ("Banesco", st.session_state.saldo_banesco),
                            ("BNC", st.session_state.saldo_bnc),
                            ("Mercantil", st.session_state.saldo_mercantil),
                            ("Banco de Venezuela (BDV)", st.session_state.saldo_venezuela),
                            ("Provincial", st.session_state.saldo_provincial),
                            ("Bancamiga", st.session_state.saldo_bancamiga),
                            ("BanPlus", st.session_state.saldo_banplus),
                            ("Banco del Tesoro", st.session_state.saldo_tesoro),
                            ("Banco Efectivo", st.session_state.saldo_efectivo),
                            ("Banco Binance", st.session_state.saldo_binance)
                        ])

                        fila_r = 9
                        for banco_n, saldo_v in bancos_data:
                            cell_b = hoja_resumen.cell(row=fila_r, column=2, value=banco_n)
                            cell_b.border = borde_fino
                            cell_b.alignment = alineacion_izquierda
                        
                            cell_s = hoja_resumen.cell(row=fila_r, column=3, value=saldo_v)
                            cell_s.border = borde_fino
                            cell_s.number_format = '#,##0.00'
                            cell_s.alignment = alineacion_derecha
                        
                            usd_v = saldo_v / tasa_dia if tasa_dia > 0 else 0.0
                            cell_u = hoja_resumen.cell(row=fila_r, column=4, value=usd_v)
                            cell_u.border = borde_fino
                            cell_u.number_format = '$#,##0.00'
                            cell_u.alignment = alineacion_derecha
                        
                            if fila_r % 2 == 0:
                                for col in range(2, 5):
                                    hoja_resumen.cell(row=fila_r, column=col).fill = gris_claro
                            fila_r += 1

                        # Totales
                        cell_total_lbl = hoja_resumen.cell(row=fila_r, column=2, value="TOTAL CONSOLIDADO")
                        cell_total_lbl.font = negro_bold
                        cell_total_lbl.border = borde_fino
                        cell_total_lbl.fill = verde_claro
                    
                        cell_total_ves = hoja_resumen.cell(row=fila_r, column=3, value=total_ves)
                        cell_total_ves.font = negro_bold
                        cell_total_ves.border = borde_fino
                        cell_total_ves.number_format = '#,##0.00'
                        cell_total_ves.fill = verde_claro
                        cell_total_ves.alignment = alineacion_derecha
                    
                        cell_total_usd = hoja_resumen.cell(row=fila_r, column=4, value=total_usd)
                        cell_total_usd.font = negro_bold
                        cell_total_usd.border = borde_fino
                        cell_total_usd.number_format = '$#,##0.00'
                        cell_total_usd.fill = verde_claro
                        cell_total_usd.alignment = alineacion_derecha

                        # Total Ingresos Archivos
                        fila_r += 1
                        cell_total_ing_lbl = hoja_resumen.cell(row=fila_r, column=2, value="TOTAL INGRESOS ARCHIVOS")
                        cell_total_ing_lbl.font = negro_bold
                        cell_total_ing_lbl.border = borde_fino
                        cell_total_ing_lbl.fill = amarillo
                    
                        tot_ing_ves_export = st.session_state.get("total_ingresos_consolidado", 0.0)
                        cell_total_ing_ves = hoja_resumen.cell(row=fila_r, column=3, value=tot_ing_ves_export)
                        cell_total_ing_ves.font = negro_bold
                        cell_total_ing_ves.border = borde_fino
                        cell_total_ing_ves.number_format = '#,##0.00'
                        cell_total_ing_ves.fill = amarillo
                        cell_total_ing_ves.alignment = alineacion_derecha
                    
                        tot_ing_usd_export = tot_ing_ves_export / tasa_dia if tasa_dia > 0 else 0.0
                        cell_total_ing_usd = hoja_resumen.cell(row=fila_r, column=4, value=tot_ing_usd_export)
                        cell_total_ing_usd.font = negro_bold
                        cell_total_ing_usd.border = borde_fino
                        cell_total_ing_usd.number_format = '$#,##0.00'
                        cell_total_ing_usd.fill = amarillo
                        cell_total_ing_usd.alignment = alineacion_derecha

                        # Total Egresos iPago
                        fila_r += 1
                        cell_total_egr_lbl = hoja_resumen.cell(row=fila_r, column=2, value="TOTAL EGRESOS IPAGO")
                        cell_total_egr_lbl.font = negro_bold
                        cell_total_egr_lbl.border = borde_fino
                        cell_total_egr_lbl.fill = amarillo
                    
                        tot_egr_ves_export = st.session_state.get("total_egresos_ipago_ves", 0.0)
                        cell_total_egr_ves = hoja_resumen.cell(row=fila_r, column=3, value=tot_egr_ves_export)
                        cell_total_egr_ves.font = negro_bold
                        cell_total_egr_ves.border = borde_fino
                        cell_total_egr_ves.number_format = '#,##0.00'
                        cell_total_egr_ves.fill = amarillo
                        cell_total_egr_ves.alignment = alineacion_derecha
                    
                        tot_egr_usd_export = tot_egr_ves_export / tasa_dia if tasa_dia > 0 else 0.0
                        cell_total_egr_usd = hoja_resumen.cell(row=fila_r, column=4, value=tot_egr_usd_export)
                        cell_total_egr_usd.font = negro_bold
                        cell_total_egr_usd.border = borde_fino
                        cell_total_egr_usd.number_format = '$#,##0.00'
                        cell_total_egr_usd.fill = amarillo
                        cell_total_egr_usd.alignment = alineacion_derecha

                        for columna in hoja_resumen.columns:
                            max_length = 0
                            try:
                                columna_letra = columna[0].column_letter
                            except:
                                continue
                            for cell in columna:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 5, 50)
                            hoja_resumen.column_dimensions[columna_letra].width = adjusted_width

                        # -----------------------------------------------------
                        # PESTAÑA: DETALLE DE CONCILIACIÓN (REPORTE)
                        # -----------------------------------------------------
                        hoja = workbook.create_sheet(title="REPORTE")
                        if "Sheet" in workbook.sheetnames:
                            workbook.remove(workbook["Sheet"])

                        try:
                            logo = Image("LOGO.jpeg")
                            logo.width = 130
                            logo.height = 130
                            hoja.add_image(logo, "A1")
                        except:
                            pass

                        hoja.merge_cells("C7:H7")
                        hoja["C7"] = "REPORTE CONSOLIDADO DE CONCILIACIÓN MULTIBANCO"
                        hoja["C7"].font = Font(bold=True, size=14)
                        hoja["C7"].alignment = alineacion_centro

                        def crear_tabla(titulo, dataframe, fila_inicio, color_total):
                            hoja.merge_cells(start_row=fila_inicio, start_column=1, end_row=fila_inicio, end_column=10)
                            titulo_cell = hoja.cell(row=fila_inicio, column=1)
                            titulo_cell.value = titulo
                            titulo_cell.fill = rojo
                            titulo_cell.font = blanco_bold
                            titulo_cell.alignment = alineacion_centro

                            headers = [
                                "FECHA", "REFERENCIA", "DESCRIPCIÓN", "DESCRIPCIÓN ORIGINAL",
                                "MONTO BS", "TASA BCV", "MONTO USD", 
                                "PROVEEDOR (iPago)", "TIPO EGRESO (iPago)", "TIPO PAGO (iPago)"
                            ]
                            fila_header = fila_inicio + 1

                            for col_num, header in enumerate(headers, 1):
                                cell = hoja.cell(row=fila_header, column=col_num)
                                cell.value = header
                                cell.fill = rojo
                                cell.font = blanco_bold
                                cell.border = borde_fino
                                cell.alignment = alineacion_centro

                            fila_data = fila_header + 1

                            for _, row in dataframe.iterrows():
                                hoja.cell(row=fila_data, column=1).value = row.get("FECHA", "")
                                hoja.cell(row=fila_data, column=2).value = row.get("REFERENCIA", "")
                                hoja.cell(row=fila_data, column=3).value = row.get("DESCRIPCIÓN", "")
                                hoja.cell(row=fila_data, column=4).value = row.get("DESCRIPCION_ORIGINAL", "")
                                hoja.cell(row=fila_data, column=5).value = row.get("MONTO BS", 0)
                                hoja.cell(row=fila_data, column=6).value = row.get("TASA BCV", 0)
                                hoja.cell(row=fila_data, column=7).value = row.get("MONTO USD", 0)
                                hoja.cell(row=fila_data, column=8).value = row.get("PROVEEDOR_IPAGO", row.get("STATUS", ""))
                                hoja.cell(row=fila_data, column=9).value = row.get("OBSERVACIÓN", "")
                                hoja.cell(row=fila_data, column=10).value = row.get("TIPO_PAGO", "")

                                hoja.cell(row=fila_data, column=5).number_format = '#,##0.00'
                                hoja.cell(row=fila_data, column=6).number_format = '#,##0.0000'
                                hoja.cell(row=fila_data, column=7).number_format = '$#,##0.00'

                                for col in range(1, 11):
                                     hoja.cell(row=fila_data, column=col).border = borde_fino
                                fila_data += 1

                            total_cell = hoja.cell(row=fila_data, column=4)
                            total_cell.value = f"TOTAL {titulo}"
                            total_cell.font = Font(bold=True)

                            total_bs_cell = hoja.cell(row=fila_data, column=5)
                            total_bs_cell.value = dataframe["MONTO BS"].sum() if not dataframe.empty else 0
                            total_bs_cell.number_format = '#,##0.00'
                            total_bs_cell.fill = color_total

                            monto_total = hoja.cell(row=fila_data, column=7)
                            monto_total.value = dataframe["MONTO USD"].sum() if not dataframe.empty else 0
                            monto_total.number_format = '$#,##0.00'
                            monto_total.fill = color_total

                            return fila_data + 4

                        fila_actual = 10
                        if not df_ingresos.empty:
                            fila_actual = crear_tabla("INGRESOS", df_ingresos, fila_actual, verde_claro)
                        if not df_egresos.empty:
                            fila_actual = crear_tabla("EGRESOS", df_egresos, fila_actual, amarillo)
                        if not df_comisiones.empty:
                            fila_actual = crear_tabla("COMISIONES", df_comisiones, fila_actual, amarillo)

                        for columna in hoja.columns:
                            max_length = 0
                            try:
                                columna_letra = columna[0].column_letter
                            except:
                                continue
                            for cell in columna:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 5, 50)
                            hoja.column_dimensions[columna_letra].width = adjusted_width

                    output.seek(0)

                    st.download_button(
                        label="📥 Descargar Excel Clasificado Consolidado (BCV + iPago)",
                        data=output.getvalue(),
                        file_name=f"cierre_consolidado_{fecha_inicio}_{fecha_fin}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                
                    with st.expander("📊 Tasas BCV utilizadas en el proceso"):
                        todas_tasas = {}
                        for registro in ingresos + egresos + comisiones:
                            fecha_r = registro["FECHA"]
                            tasa_r = registro["TASA BCV"]
                            if fecha_r not in todas_tasas:
                                todas_tasas[fecha_r] = tasa_r
                        if todas_tasas:
                            df_tasas = pd.DataFrame([
                                {"FECHA": f, "TASA BCV": t} 
                                for f, t in todas_tasas.items()
                            ]).sort_values("FECHA")
                            st.dataframe(df_tasas, use_container_width=True)

    else:
        st.session_state.total_ingresos_consolidado = 0.0
        st.session_state.total_egresos_ipago_ves = 0.0
        st.markdown("""
        ### 👋 Conciliador Bancario Inteligente Multibanco
    
        Carga los archivos de tus cuentas bancarias y el archivo maestro de iPago en el menú de la izquierda para comenzar el proceso de conciliación automatizado.
    
        **Características:**
        - Soporte simultáneo para múltiples cuentas.
        - Cálculo automático de saldo consolidado en Bolívares (VES) y Dólares (USD).
        - Cruce inteligente y trazabilidad con iPago.
        - Generación de reportes de cierre en formato Excel profesional.
        """)
else:
    # ---------------------------------------------------------
    # FLUX 2: CRUCE DE INFORMACIÓN (MONOBANCO)
    # ---------------------------------------------------------
    # =========================================================

    df_ipago = None

    if archivo:
        st.info(f"📄 Archivo: **{archivo.name}** - {archivo.size/1024:.1f} KB")

        try:
            # =========================================================
            # 🔥 DETECCIÓN DE BANCO - NUEVO ORDEN
            # =========================================================
        
            # 1. Detectar SIEMPRE por el nombre primero
            banco = mono_detectar_banco_por_nombre(archivo.name)
        
            # 2. Solo si no se reconoce (queda como "mercantil"), intentar por contenido
            if banco == "mercantil":
                banco_contenido = mono_detectar_banco_por_contenido(archivo)
                if banco_contenido:
                    banco = banco_contenido
        
            st.success(f"🏦 **Banco detectado:** {banco.upper()}")
        
            if banco == "mercantil":
                df_original = mono_leer_excel_sin_encabezados(archivo)
            
            elif banco == "banesco":
                try:
                    nombre = archivo.name.lower()
                    if nombre.endswith(".xlsx") or nombre.endswith(".xlsm"):
                        df_raw = pd.read_excel(archivo, engine="openpyxl", header=None)
                    else:
                        df_raw = pd.read_html(archivo)[0]
                    df_normalizado = mono_procesar_banesco(df_raw)
                    df_original = mono_convertir_a_formato_mercantil(df_normalizado, banco)
                except Exception as e:
                    st.error(f"Error leyendo Banesco: {str(e)}")
                    st.stop()
            
            elif banco == "tesoro":
                try:
                    df_raw = pd.read_excel(archivo, engine="openpyxl")
                    df_normalizado = mono_procesar_tesoro(df_raw)
                    df_original = mono_convertir_a_formato_mercantil(df_normalizado, banco)
                except Exception as e:
                    st.error(f"Error leyendo Tesoro: {str(e)}")
                    st.stop()
            
            elif banco == "bancamiga":
                try:
                    # 🔥 CARGA DE BANCAMIGA
                    nombre = archivo.name.lower()
                
                    if nombre.endswith(".xlsx") or nombre.endswith(".xlsm"):
                        df_raw = pd.read_excel(archivo, engine="openpyxl", header=None)
                    else:
                        try:
                            # Intentar leer como Excel .xls real
                            df_raw = pd.read_excel(archivo, header=None)
                        except Exception:
                            # Si realmente es HTML disfrazado de .xls
                            archivo.seek(0)
                            df_raw = pd.read_html(archivo, decimal=',', thousands='.')[0]
                    
                    if isinstance(df_raw.columns, pd.MultiIndex):
                        df_raw.columns = df_raw.columns.get_level_values(-1)
                
                    df_normalizado = mono_procesar_bancamiga(df_raw)
                    if df_normalizado.empty:
                        st.error("No se pudieron procesar los datos de Bancamiga.")
                        st.stop()
                    df_original = mono_convertir_a_formato_mercantil(df_normalizado, banco)
                except Exception as e:
                    st.error(f"Error leyendo Bancamiga: {str(e)}")
                    st.stop()
            
            elif banco == "provincial":
                try:
                    df_raw = mono_leer_excel_sin_encabezados(archivo)
                    df_normalizado = mono_procesar_provincial(df_raw)
                    if df_normalizado.empty:
                        st.error("No se pudieron procesar los datos de Provincial.")
                        st.stop()
                    df_original = mono_convertir_a_formato_mercantil(df_normalizado, banco)
                except Exception as e:
                    st.error(f"Error leyendo Provincial: {str(e)}")
                    st.stop()
            
            elif banco == "venezuela":
                # 🔥 LEER SIN ENCABEZADOS (igual que Mercantil)
                df_raw = mono_leer_excel_sin_encabezados(archivo)
                df_normalizado = mono_procesar_venezuela_simple(df_raw)
                if df_normalizado.empty:
                    st.stop()
            
                # Venezuela trabaja directamente con el dataframe normalizado
                df_original = mono_convertir_venezuela_a_formato_mercantil(df_normalizado)
            
                # Fechas para Venezuela
                fechas_convertidas = pd.to_datetime(
                    df_normalizado["FECHA"],
                    dayfirst=True,
                    errors="coerce"
                )
            
            elif banco == "banplus":
                try:
                    df_raw = mono_leer_excel_sin_encabezados(archivo)
                    df_normalizado = mono_procesar_banplus(df_raw)
                    if df_normalizado.empty:
                        st.error("No se pudieron procesar los datos de Banplus.")
                        st.stop()
                    df_original = mono_convertir_a_formato_mercantil(df_normalizado, banco)
                except Exception as e:
                    st.error(f"Error leyendo Banplus: {str(e)}")
                    st.stop()
            
            elif banco == "bnc":
                df_raw = leer_excel_con_encabezados(archivo)
                df_normalizado = mono_procesar_bnc(df_raw)
                df_original = mono_convertir_a_formato_mercantil(df_normalizado, banco)
            
            else:
                df_raw = leer_excel_con_encabezados(archivo)
                df_original = mono_convertir_a_formato_mercantil(df_raw, banco)
            
            if df_original.empty:
                st.error("No se detectaron movimientos para procesar.")
                st.stop()

            try:
                if banco == "mercantil":
                    fechas_convertidas = pd.to_datetime(
                        df_original[3].astype(str).str.zfill(8),
                        format="%d%m%Y",
                        errors="coerce"
                    )
                elif banco == "venezuela":
                    # Ya tenemos fechas_convertidas definidas arriba
                    pass
                else:
                    fechas_convertidas = pd.to_datetime(
                        df_original.iloc[:, 3],
                        errors="coerce",
                        dayfirst=True
                    )

                fecha_inicio_dt = pd.to_datetime(fecha_inicio)
                fecha_fin_dt = pd.to_datetime(fecha_fin)

                # Ajustar automáticamente el rango si las fechas del archivo están fuera del rango seleccionado
                if not fechas_convertidas.empty:
                    min_file_date = fechas_convertidas.min()
                    max_file_date = fechas_convertidas.max()
                    if pd.notna(min_file_date) and pd.notna(max_file_date):
                        if fecha_inicio_dt > min_file_date or fecha_fin_dt < max_file_date:
                            fecha_inicio_dt = min_file_date
                            fecha_fin_dt = max_file_date
                            st.info(f"💡 **Rango de fechas ajustado automáticamente** al contenido del archivo: {min_file_date.strftime('%d/%m/%Y')} al {max_file_date.strftime('%d/%m/%Y')}")

                # Aplicar filtro según el banco
                if banco == "venezuela":
                    # Usar el dataframe original para el filtro
                    mask = (fechas_convertidas >= fecha_inicio_dt) & (fechas_convertidas <= fecha_fin_dt)
                    # Filtrar df_original usando la máscara
                    df_original = df_original[mask]
                else:
                    df_original = df_original[
                        (fechas_convertidas >= fecha_inicio_dt) & 
                        (fechas_convertidas <= fecha_fin_dt)
                    ]
            
                st.success(f"Filtro de fechas aplicado: {fecha_inicio_dt.strftime('%d/%m/%Y')} a {fecha_fin_dt.strftime('%d/%m/%Y')}")
            except Exception as e:
                st.warning(f"Error filtrando fechas: {e}")
            
            if df_original.empty or len(df_original) == 0:
                st.error("❌ No se encontraron movimientos válidos en el rango de fechas.")
                st.stop()

            with st.expander("👁️ Vista previa archivo original"):
                st.dataframe(df_original.head(20), use_container_width=True)

            # =========================================================
            # LEER ARCHIVO IPAGO
            # =========================================================
            if archivo_ipago:
                try:
                    df_ipago = pd.read_excel(archivo_ipago, engine="openpyxl")
                    df_ipago.columns = [str(c).strip() for c in df_ipago.columns]
                    st.success(f"Archivo iPago cargado: {len(df_ipago)} registros")
                    st.dataframe(df_ipago.head())
                except Exception as e:
                    st.error(f"Error leyendo archivo iPago: {e}")

            if procesar:
                with st.spinner("Procesando archivo con tasas BCV..."):
                    if banco == "venezuela":
                        ingresos = []
                        egresos = []
                        comisiones = []

                        for _, row in df_normalizado.iterrows():
                            fecha_obj = pd.to_datetime(row["FECHA"], dayfirst=True, errors="coerce")
                            tasa = obtener_tasa_por_fecha(fecha_obj, usar_api) or 1.0

                            monto_bs = float(row["MONTO"])
                            monto_usd = calcular_usd(monto_bs, tasa)

                            registro = {
                                "FECHA": row["FECHA"],
                                "REFERENCIA": row["REFERENCIA"],
                                "DESCRIPCIÓN": row["DESCRIPCION"],
                                "MONTO BS": monto_bs,
                                "TASA BCV": tasa,
                                "MONTO USD": monto_usd,
                                "STATUS": "",
                                "OBSERVACIÓN": "",
                                "TIPO_PAGO": "",
                                "PROVEEDOR_IPAGO": "",
                                "DESCRIPCION_ORIGINAL": ""
                            }

                            tipo = str(row["TIPO"]).strip().upper()
                            descripcion = str(row["DESCRIPCION"]).strip().upper()
                            referencia = str(row["REFERENCIA"]).strip()
                        
                            # 🔥 DETECCIÓN MEJORADA DE COMISIONES PARA VENEZUELA
                            es_comision_venezuela = False
                        
                            # 🔥 REGLA 1: Detectar por descripción
                            palabras_comision_bdv = [
                                "COM PAGO OTRAS CTAS",
                                "COMISION PAGO A PROVEEDORES",
                                "COM PAGO OTR BCOS",
                                "COM PAGO OTRAS CTAS JUR NAT",
                                "COM PAGO OTRAS CTAS JUR JUR",
                                "COMISION POR TRANSFERENCIA",
                                "COMISION PAGO MOVIL",
                                "COMISIÓN PAGO MOVIL",
                                "COMISION X PAGO DE NOMINA",
                                "COMISION X PAGO DE NOMINAS",
                                "ITF",
                                "IMPUESTO A LAS TRANSACCIONES FINANCIERAS",
                                "CARGO BANCARIO",
                                "MANTENIMIENTO DE CUENTA",
                                "COMISION BANCARIA",
                                "COMISIÓN BANCARIA",
                                "CARGO POR SERVICIO",
                                "CARGO POR TRANSACCION",
                                "COMISION PAGO MOVIL COMERCIAL",
                                "COMISION X PAGO DE NOMINAS MB",
                                "DOMICILIACION J412438905",
                                "DISTRIBUIDORA GLOBAL",
                                "DOMICILIACION"
                            ]
                        
                            # Verificar si la descripción coincide con alguna comisión
                            for patron in palabras_comision_bdv:
                                if patron in descripcion:
                                    es_comision_venezuela = True
                                    break
                        
                            # 🔥 REGLA 2: Detectar por referencia (comisiones de BDV tienen referencias específicas)
                            if not es_comision_venezuela:
                                # Las comisiones de BDV suelen tener referencias que comienzan con 970 o 972 o 067
                                if referencia.startswith(("970", "972", "067")):
                                    # Verificar si la descripción contiene palabras clave de comisión
                                    if any(palabra in descripcion for palabra in ["COM", "PAGO OTRAS", "PAGO OTR", "COMISION"]):
                                        es_comision_venezuela = True
                        
                            # 🔥 REGLA 3: Si el tipo es ND y la descripción contiene "COM" es una comisión
                            if not es_comision_venezuela and tipo == "ND":
                                if "COM" in descripcion or "PAGO OTR" in descripcion:
                                    es_comision_venezuela = True
                        
                            # 🔥 REGLA 4: Comisiones específicas de BDV por el monto exacto (189.50, 36.44, etc)
                            if not es_comision_venezuela and tipo == "ND":
                                # Montos típicos de comisiones de BDV (montos pequeños)
                                if monto_bs < 1000 and ("COM" in descripcion or "PAGO OTR" in descripcion):
                                    es_comision_venezuela = True
                        
                            # Si es comisión, agregar a la lista de comisiones
                            if es_comision_venezuela:
                                comisiones.append(registro)
                            elif tipo in ["NC", "C", "CREDITO", "ABONO"]:
                                ingresos.append(registro)
                            else:
                                egresos.append(registro)
                    else:
                        # 🔥 PASAR EL NOMBRE DEL BANCO A LA FUNCIÓN
                        ingresos, egresos, comisiones = mono_procesar_archivo(df_original, usar_api, banco=banco)

                df_ingresos = pd.DataFrame(ingresos)
                df_egresos = pd.DataFrame(egresos)
                df_comisiones = pd.DataFrame(comisiones)

                # =========================================================
                # 🔥 CRUCE CON IPAGO - VERSIÓN MEJORADA (CRUCE FLEXIBLE)
                # =========================================================
                if archivo_ipago and not df_egresos.empty:
                    # Enriquecer egresos con datos de iPago
                    df_egresos = mono_enriquecer_egresos_con_ipago(df_egresos, df_ipago)
                
                    # 🔥 Separar comisiones de iPago (si las hay)
                    if "ES_COMISION" in df_egresos.columns:
                        mascara_comisiones_ipago = df_egresos["ES_COMISION"] == True
                    
                        if mascara_comisiones_ipago.any():
                            df_comisiones_extra = df_egresos[mascara_comisiones_ipago].copy()
                        
                            # Remover columnas internas
                            df_comisiones_extra = df_comisiones_extra.drop(
                                columns=["ES_COMISION", "REFERENCIA_IPAGO"], 
                                errors="ignore"
                            )
                        
                            # Agregar a comisiones existentes
                            if not df_comisiones.empty:
                                df_comisiones = pd.concat([df_comisiones, df_comisiones_extra], ignore_index=True)
                            else:
                                df_comisiones = df_comisiones_extra
                        
                            # Remover comisiones de egresos
                            df_egresos = df_egresos[~mascara_comisiones_ipago].copy()
                        
                            st.success(f"💳 Se movieron {len(df_comisiones_extra)} comisiones desde iPago a la sección de COMISIONES")
                
                    # Limpiar columnas auxiliares
                    df_egresos = df_egresos.drop(
                        columns=["ES_COMISION", "REFERENCIA_IPAGO"], 
                        errors="ignore"
                    )
                
                    st.success(f"🎯 Egresos enriquecidos con iPago: {len(df_egresos)} registros")

                # Completar columnas vacías obligatorias para el reporte en openpyxl
                for df_t in [df_ingresos, df_egresos, df_comisiones]:
                    if not df_t.empty:
                        if "STATUS" not in df_t.columns: df_t["STATUS"] = ""
                        if "OBSERVACIÓN" not in df_t.columns: df_t["OBSERVACIÓN"] = ""
                        if "TIPO_PAGO" not in df_t.columns: df_t["TIPO_PAGO"] = ""
                        if "PROVEEDOR_IPAGO" not in df_t.columns: df_t["PROVEEDOR_IPAGO"] = ""
                        if "DESCRIPCION_ORIGINAL" not in df_t.columns: df_t["DESCRIPCION_ORIGINAL"] = ""

                total_ingresos = df_ingresos["MONTO USD"].sum() if not df_ingresos.empty else 0
                total_egresos = df_egresos["MONTO USD"].sum() if not df_egresos.empty else 0
                total_comisiones = df_comisiones["MONTO USD"].sum() if not df_comisiones.empty else 0

                col1, col2, col3 = st.columns(3)

                with col1:
                    st.metric("💰 INGRESOS", len(df_ingresos), f"${total_ingresos:,.2f}")
                with col2:
                    st.metric("💸 EGRESOS", len(df_egresos), f"${total_egresos:,.2f}")
                with col3:
                    st.metric("💳 COMISIONES", len(df_comisiones), f"${total_comisiones:,.2f}")

                st.subheader("📊 Resultados")

                tab1, tab2, tab3 = st.tabs(["📈 INGRESOS", "📉 EGRESOS", "💳 COMISIONES"])

                with tab1: st.dataframe(df_ingresos, use_container_width=True)
                with tab2: st.dataframe(df_egresos, use_container_width=True)
                with tab3: st.dataframe(df_comisiones, use_container_width=True)

                # =========================================================
                # MOTOR DE REPORTES EXCEL OPENPYXL COMPLETO
                # =========================================================
                output = BytesIO()

                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    workbook = writer.book
                    hoja = workbook.create_sheet(title="REPORTE")

                    if "Sheet" in workbook.sheetnames:
                        hoja_vacia = workbook["Sheet"]
                        workbook.remove(hoja_vacia)

                    rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    verde = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                    amarillo = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    blanco = Font(color="FFFFFF", bold=True)
                    borde = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
                    centro = Alignment(horizontal="center", vertical="center")

                    try:
                        logo = Image("LOGO.jpeg")
                        logo.width = 130
                        logo.height = 130
                        hoja.add_image(logo, "A1")
                    except:
                        pass

                    hoja.merge_cells("C7:H7")
                    banco_nombre = banco.upper()
                    hoja["C7"] = f"{banco_nombre} - REPORTE DE CONCILIACIÓN"
                    hoja["C7"].font = Font(bold=True, size=14)
                    hoja["C7"].alignment = centro

                    def crear_tabla(titulo, dataframe, fila_inicio, color_total):
                        hoja.merge_cells(start_row=fila_inicio, start_column=1, end_row=fila_inicio, end_column=10)
                        titulo_cell = hoja.cell(row=fila_inicio, column=1)
                        titulo_cell.value = titulo
                        titulo_cell.fill = rojo
                        titulo_cell.font = blanco
                        titulo_cell.alignment = centro

                        # 🔥 HEADERS MEJORADOS CON DATOS DE IPAGO
                        headers = [
                            "FECHA", "REFERENCIA", "DESCRIPCIÓN", "DESCRIPCIÓN ORIGINAL",
                            "MONTO BS", "TASA BCV", "MONTO USD", 
                            "PROVEEDOR (iPago)", "TIPO EGRESO (iPago)", "TIPO PAGO (iPago)"
                        ]
                        fila_header = fila_inicio + 1

                        for col_num, header in enumerate(headers, 1):
                            cell = hoja.cell(row=fila_header, column=col_num)
                            cell.value = header
                            cell.fill = rojo
                            cell.font = blanco
                            cell.border = borde
                            cell.alignment = centro

                        fila_data = fila_header + 1

                        for _, row in dataframe.iterrows():
                            hoja.cell(row=fila_data, column=1).value = row.get("FECHA", "")
                            hoja.cell(row=fila_data, column=2).value = row.get("REFERENCIA", "")
                            hoja.cell(row=fila_data, column=3).value = row.get("DESCRIPCIÓN", "")
                            hoja.cell(row=fila_data, column=4).value = row.get("DESCRIPCION_ORIGINAL", "")
                            hoja.cell(row=fila_data, column=5).value = row.get("MONTO BS", 0)
                            hoja.cell(row=fila_data, column=6).value = row.get("TASA BCV", 0)
                            hoja.cell(row=fila_data, column=7).value = row.get("MONTO USD", 0)
                            hoja.cell(row=fila_data, column=8).value = row.get("PROVEEDOR_IPAGO", row.get("STATUS", ""))
                            hoja.cell(row=fila_data, column=9).value = row.get("OBSERVACIÓN", "")
                            hoja.cell(row=fila_data, column=10).value = row.get("TIPO_PAGO", "")

                            hoja.cell(row=fila_data, column=5).number_format = '#,##0.00'
                            hoja.cell(row=fila_data, column=6).number_format = '#,##0.0000'
                            hoja.cell(row=fila_data, column=7).number_format = '$#,##0.00'

                            for col in range(1, 11):
                                hoja.cell(row=fila_data, column=col).border = borde

                            fila_data += 1

                        total_cell = hoja.cell(row=fila_data, column=4)
                        total_cell.value = f"TOTAL {titulo}"
                        total_cell.font = Font(bold=True)

                        total_bs_cell = hoja.cell(row=fila_data, column=5)
                        total_bs_cell.value = dataframe["MONTO BS"].sum() if not dataframe.empty else 0
                        total_bs_cell.number_format = '#,##0.00'
                        total_bs_cell.fill = color_total

                        monto_total = hoja.cell(row=fila_data, column=7)
                        monto_total.value = dataframe["MONTO USD"].sum() if not dataframe.empty else 0
                        monto_total.number_format = '$#,##0.00'
                        monto_total.fill = color_total

                        return fila_data + 4

                    fila_actual = 10

                    if not df_ingresos.empty:
                        fila_actual = crear_tabla("INGRESOS", df_ingresos, fila_actual, verde)

                    if not df_egresos.empty:
                        fila_actual = crear_tabla("EGRESOS", df_egresos, fila_actual, amarillo)

                    if not df_comisiones.empty:
                        fila_actual = crear_tabla("COMISIONES", df_comisiones, fila_actual, amarillo)

                    for columna in hoja.columns:
                        max_length = 0
                        try:
                            columna_letra = columna[0].column_letter
                        except:
                            continue

                        for cell in columna:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass

                        adjusted_width = min(max_length + 5, 50)
                        hoja.column_dimensions[columna_letra].width = adjusted_width

                output.seek(0)

                st.download_button(
                    label="📥 Descargar Excel Clasificado (con Tasas BCV e iPago)",
                    data=output.getvalue(),
                    file_name=f"balance_{banco}_{fecha_inicio}_{fecha_fin}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
                with st.expander("📊 Tasas BCV utilizadas"):
                    todas_tasas = {}
                    for registro in ingresos + egresos + comisiones:
                        fecha = registro["FECHA"]
                        tasa = registro["TASA BCV"]
                        if fecha not in todas_tasas:
                            todas_tasas[fecha] = tasa
                
                    if todas_tasas:
                        df_tasas = pd.DataFrame([
                            {"FECHA": f, "TASA BCV": t} 
                            for f, t in todas_tasas.items()
                        ]).sort_values("FECHA")
                        st.dataframe(df_tasas, use_container_width=True)

        except Exception as e:
            st.error(f"❌ Error general: {str(e)}")
            st.error("Detalles del error para depuración:")
            st.code(str(e))

    else:
        st.markdown("""
        ### 👋 Clasificador Bancario Inteligente Multi-Banco

        ## FUNCIONES
        ✅ **Bancos soportados:** Mercantil, Banco de Venezuela, Banesco, Provincial, BNC, Tesoro, Bancamiga.
        ✅ Clasifica automáticamente: Ingresos (NC, C, CREDITO, ABONO), Egresos (ND, D, DEBITO, DEBIT), Comisiones.
        ✅ **NUEVO:** Cruce inteligente con iPago usando REFERENCIA + DESCRIPCIÓN.
        ✅ **NUEVO:** Exporta con datos completos de iPago: Proveedor, Tipo de Egreso, Tipo de Pago.
        ✅ **NUEVO:** Conserva la descripción original del banco y la reemplaza con la de iPago cuando hay coincidencia.
        ✅ Calcula USD con tasa BCV real por fecha.
        ✅ Exporta reporte profesional con todas las columnas.
        """)
